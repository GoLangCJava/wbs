# -*- coding: utf-8 -*-
"""功能拆解估算（v3）：26项功能下钻子任务 + 非功能与工程支撑区块，
按 BA / 开发A(Wade) / 开发B(DevB) / 测试 / SLC文档 直估人天。合并单元格版式。
用法：python3 func_breakdown.py
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PATH = "/home/user/wbs/RIMS项目WBS工作分解_AI Native版.xlsx"
wb = openpyxl.load_workbook(PATH)

# ---------- 功能子任务：(功能编号, [(子任务, 要点, BA, Wade, DevB, Test, SLC), ...]) ----------
B = [
("F01-01", [
 ("原系统基本信息录入", "来源系统档案：系统编码/名称/业务域/责任人/退役批次；录入与维护界面+存储API", .5, 1.5, .5, .5, .25),
 ("源库DB连接配置", "DB类型/连接串录入，账号密码Key Vault加密存储，连通性测试", .25, 1.5, 1, .5, .25),
 ("同步表勾选", "拉取源表清单、勾选同步范围、数据量预估与全选过滤", .25, 1.5, 1, .5, .25),
 ("保留期限与迁移参数配置", "保留策略绑定、迁移批次命名、并发/限速参数", .5, 1, .5, .25, .25),
 ("存储落地实现（Iceberg/UC）", "磁盘目录布局、Iceberg表属性与分区策略、UC目录/Schema/授权配置化", .25, 0, 2, .5, .5),
 ("SeaTunnel批量导入作业", "作业配置生成器、分片/断点续传/限速执行", 0, .5, 2, 1, .5),
 ("导入执行与监控", "任务状态/进度/失败重试/日志查看", 0, 1.5, .5, .5, .25),
]),
("F01-02", [
 ("文件来源登记与上传通道", "在线上传/共享路径挂载/离线介质登记", .25, 2, 1, .5, .25),
 ("大文件分片上传与断点续传", "前端分片上传、后端合并与校验", 0, 1.5, 1.5, 1, .25),
 ("文件包解析", "zip/tar/数据库备份包识别与展开、原始属性保留", 0, 0, 1.5, .75, .25),
 ("附件元数据与Blob存储布局", "附件索引表、路径规范、与结构化记录关联", .25, 0, 1.5, .5, .25),
]),
("F01-03", [
 ("字段映射规则配置", "源字段→标准字段映射界面、自动映射建议", .5, 2, .5, .5, .25),
 ("格式转换与清洗规则", "类型转换/编码/默认值/脏数据处理策略", .5, 0, 1.5, .75, .25),
 ("校验规则配置", "必填/唯一/范围/自定义表达式校验", 0, 1, 1, .5, .25),
 ("规则版本与试跑预览", "规则版本管理、样本试跑与结果预览", 0, .5, 1, .5, .25),
]),
("F01-04", [
 ("基准采集策略定义", "行数/校验和/SHA-256抽样口径", .5, 0, 1, .25, .25),
 ("基准采集作业与存储", "源端基准快照生成、基准库表", 0, 0, 1.5, .5, .25),
 ("基准报告查询展示", "基准结果展示与导出", 0, .5, .25, .25, .25),
]),
("F01-05", [
 ("比对引擎", "RAW/CURATED逐层行数与校验和比对、抽样SHA-256", 0, 0, 2, .75, .25),
 ("差异报告与处理流程", "差异清单、重迁/豁免记录", .25, 1, .5, .5, .25),
]),
("F02-01", [
 ("元数据字段体系定义", "字段/类型/必填/密级/保留等级字典", 1.5, 0, .5, .25, .5),
 ("元数据核心表与API", "核心元数据表实现、CRUD API、Migration", 0, 0, 2, .5, .25),
 ("元数据驱动配置机制", "查询配置模型、动态表单渲染协议", .25, 2, .5, .5, .25),
]),
("F02-02", [
 ("绑定规则与批次属性", "源编码引用、退役项目/迁移批次属性规则", .5, 0, .75, .25, .25),
 ("接入时自动绑定", "导入/上传时来源属性自动写入与校验", 0, 0, 1, .5, .25),
 ("按来源检索与统计", "来源维度筛选、统计报表", 0, 1, .5, .5, .25),
]),
("F03-01", [
 ("保留策略配置", "策略规则、继承关系、生效范围", .75, 1, .5, .5, .25),
 ("处置引擎", "到期扫描、豁免检查、DROP PARTITION、VACUUM物理清除", 0, 0, 2.5, 1, .5),
 ("处置计划审批与台账", "预演清单、审批流、处置留痕台账", .25, 1.5, .5, .5, .25),
]),
("F03-02", [
 ("Hold申请与审批流程", "涉诉/审计/调查场景、审批链定义", .75, 1, .25, .5, .25),
 ("Hold生效与解除实现", "记录级/表级冻结、处置豁免、操作留痕", 0, 0, 1.5, .75, .25),
 ("Hold清单管理与提醒", "在Hold清单、解除预警", 0, .5, .25, .25, .25),
]),
("F04-01", [
 ("加密策略定义", "SSE+CMK方案、PII字段清单与字段级加密规则", .5, 0, .75, .25, .25),
 ("存储层加密落地【原生】", "ADLS SSE+CMK为存储账户原生选项：启用/托管密钥绑定/轮换/传输加密验证（配置为主）", 0, 0, .5, .5, .25),
 ("PII字段级加密服务", "应用层加解密、密文索引策略、传输加密验证", 0, 0, 1.5, .75, .25),
]),
("F04-02", [
 ("密钥全流程管理【原生】", "Key Vault原生服务：创建/保管/轮换/吊销/权限=Key Vault+RBAC配置；开发仅应用侧Managed Identity集成与密钥引用", .25, 0, .5, .5, .25),
 ("密钥操作审计与告警【原生】", "Key Vault诊断日志+Azure Monitor原生告警，开发仅对接审计页", 0, .25, .25, .5, .25),
 ("密钥状态展示（不自研管理界面）【原生】", "密钥操作走Azure门户/CLI+RBAC，不自研管理界面；系统内仅密钥状态只读展示", 0, .25, 0, .25, .25),
]),
("F04-03", [
 ("备份策略配置", "备份周期/范围/保留期策略", .5, .75, .5, .25, .25),
 ("备份作业与副本创建【原生】", "SQL DB PITR/ADLS生命周期与副本均为原生策略：仅策略配置与脚本化启用", 0, 0, .5, .5, .25),
 ("备份监控与恢复操作【原生】", "Azure Monitor原生备份告警+恢复流程验证演练", 0, .25, .25, .5, .25),
]),
("F05-01", [
 ("检索配置模型", "可检索字段、分词策略、OCR文本索引策略", .5, .5, 1, .25, .25),
 ("检索页动态表单", "元数据驱动渲染、关键词与组合查询", 0, 2, .25, .5, .25),
 ("查询代理与SERVE层对接", "查询服务、索引优化、结果高亮", 0, 0, 1.25, .75, .25),
]),
("F05-02", [
 ("来源系统筛选组件", "来源系统下拉/搜索/多选", 0, .75, 0, .25, .25),
 ("来源维度查询下推", "sys_id过滤、分区裁剪", 0, 0, .5, .25, .25),
]),
("F05-03", [
 ("时间维度筛选组件", "创建/修改/业务发生时间范围选择", 0, .75, 0, .25, .25),
 ("时间分区下推与索引", "分区裁剪、时间索引优化", 0, 0, .5, .25, .25),
]),
("F05-04", [
 ("列表展示与分页排序", "关键字段展示、分批加载、虚拟滚动", 0, 2, 0, .5, .25),
 ("大批量结果性能", "异步计数、游标分页、缓存", 0, .25, .75, .5, .25),
]),
("F05-05", [
 ("预览服务", "格式识别、图片/PDF/Office转换", 0, 0, 1.5, .5, .25),
 ("SAS直连与流式加载", "临时SAS链接、流式传输、权限校验", 0, .5, 1, .5, .25),
 ("前端预览组件", "格式适配、缩放、失败降级下载", 0, 1.5, 0, .25, .25),
]),
("F06-01", [
 ("埋点规范与日志模型", "操作对象/用户/时间/结果/IP模型定义", .25, 0, .5, .25, .25),
 ("操作拦截与落库", "归档/导入/上传统一拦截器、异步写入", 0, 0, 1, .5, .25),
 ("审计查询界面", "多维筛选与导出（覆盖归档/查询/导出三类日志）", 0, 1.5, .25, .5, .25),
]),
("F06-02", [
 ("查询/预览行为埋点", "查询条件脱敏记录、预览行为记录", 0, 0, .75, .5, .25),
 ("异步批量写入", "缓冲批量写入、不阻塞查询主链路", 0, 0, .5, .25, .25),
]),
("F06-03", [
 ("导出行为审计", "导出范围/条数/导出人/IP/下载链接追踪", 0, 0, .75, .5, .25),
 ("大批量导出告警", "阈值告警、审批联动", 0, .5, .25, .25, .25),
]),
("F08-01", [
 ("RBAC模型与权限矩阵", "角色/菜单/数据范围/记录类型/操作权限定义", 1, 0, .5, .25, .5),
 ("角色管理界面与API", "角色CRUD、权限分配界面与接口", 0, 2, .5, .5, .25),
 ("权限校验与UC授权映射", "菜单/接口级校验中间件、UC权限同步", 0, 0, 1, .75, .25),
]),
("F08-02", [
 ("账号CRUD与生命周期", "创建/启用禁用/密码重置/有效期", .25, 1, .75, .5, .25),
 ("账号与角色绑定", "分配、批量操作、人员交接", 0, .5, .25, .25, .25),
]),
("F08-03", [
 ("Entra ID对接配置【原生】", "应用注册/回调/claims为Entra控制台配置，开发仅环境参数化", .25, .5, .5, .5, .25),
 ("SSO登录链路【中间件】", "MSAL/OIDC标准库承载认证流，自研仅登录页/会话与定制逻辑", 0, 1.25, .5, .75, .25),
 ("认证异常场景处理", "认证失败、账号锁定、多账号冲突", 0, .5, 0, .5, .25),
]),
("F08-04", [
 ("身份映射规则", "企业账号↔RIMS用户映射、自动建号规则", .5, 0, .5, .25, .25),
 ("首登自动关联与角色赋予", "默认角色、待分配队列", 0, .25, .75, .5, .25),
]),
("F08-05", [
 ("会话与超时策略", "超时控制、登出、单点登出", .25, 1, .25, .5, .25),
 ("异常访问控制", "账号锁定、认证失效限制访问、IP策略", 0, 0, .75, .5, .25),
]),
]

# ---------- 非功能与工程支撑区块：(组编码, 组名, [子任务...]) ----------
N = [
("N1", "【非功能】工程初始化与开发基座", [
 ("项目初始化", "代码仓库/分支策略/开发环境统一/AI工具账号开通", 0, 1, .25, .25, 0),
 ("前端工程初始化", "React骨架/路由/布局/组件库/状态管理/构建链", 0, 1.5, 0, .25, 0),
 ("后端工程初始化", ".NET Clean Architecture/EF Core/Docker Compose本地环境", 0, 0, 1.5, .25, 0),
 ("AI开发工具链与编码规范", "提示词模板库/人机协作规约/ADR规范", 0, 1, .25, 0, .25),
 ("四道安全门禁CI/CD", "构建/镜像/SAST·依赖·密钥扫描/越权用例门禁，流水线落地与拦截演示", 0, 3, .5, .5, .25),
]),
("N2", "【非功能】代码质量与静态扫描治理", [
 ("前端ESLint/Prettier规约", "规则集接入CI、存量告警清零、格式统一", 0, 1.5, 0, 0, 0),
 ("后端SonarLint规约", "代码异味/重复块/复杂度阈值治理清零", 0, 0, 1.5, 0, 0),
 ("SAST安全告警修复", "硬编码密钥/注入/越权路径等静态安全缺陷修复", 0, 1, 1, .25, 0),
 ("依赖漏洞治理（SCA）", "依赖版本锁定、漏洞升级、SBOM产出", 0, .5, .5, 0, .25),
 ("技术债偿还与外部评审落实", "每周固定还债窗口、关键模块重构、评审意见闭环", 0, 1, 1, .25, 0),
]),
("N3", "【非功能】全局非功能实现与优化", [
 ("前端性能优化", "懒加载/虚拟滚动/缓存/首屏指标（对照NFR基线）", 0, 1.5, 0, .5, 0),
 ("查询与数据性能", "分区裁剪/Z-Order/Compaction/小文件治理", 0, 0, 1.5, .5, 0),
 ("统一鉴权与数据掩码中间件", "菜单/API/行级三层校验、掩码旁路防护", .25, 1, 1, .75, 0),
 ("可观测性（日志/监控/告警）【原生】", "App Insights/Log Analytics/Azure Monitor原生：开发仅SDK插桩、作业监控对接与仪表盘", 0, .5, .75, .25, 0),
 ("容量与增长验证", "600~800表规模扩容压测与容量报告", 0, 0, 1, .5, .25),
]),
("N4", "【非功能】部署与运维支撑", [
 ("Azure DevOps连接K8s与CI/CD部署流水线", "Service Connection/命名空间RBAC授权、构建→镜像→部署K8s流水线（测试/生产两套）、生产环境审批门禁", 0, 1, .25, 0, .25),
 ("K8s发布清单与环境配置", "原生清单Deployment/Service/Ingress/ConfigMap+Secret（不用Helm）、测试/生产两套配置、镜像标签策略/发布Runbook", 0, 1, .25, 0, 0),
 ("测试环境先期部署", "早期先部署测试环境（供联调/集成/系统测试）、随迭代由流水线自动更新", 0, 0, .5, .25, 0),
 ("部署预演与回退预案", "预生产完整演练、回退验证", 0, 0, .5, 0, 0),
 ("上线值守与运维交接", "上线窗口值守、运维交接与培训材料", 0, 1, .25, 0, .25),
]),
("N5", "【非功能】开发侧文档（架构/设计/接口）", [
 ("架构与总体设计文档", "架构图/部署视图/ADR汇编成文（Wade提供）", 0, 1.5, 0, 0, 0),
 ("应用接口与模块设计文档", "OpenAPI导出+模块设计说明（Wade提供）", 0, 1, 0, 0, 0),
 ("数据设计与ETL/作业文档", "分层设计/数据字典/ETL与作业说明（DevB提供）", 0, 0, 2, 0, 0),
]),
]

# ---------- SLC按功能/组级预估（不细到子任务；开发侧文档见N5） ----------
SLC_BY_GRP = {
 "F01-01": 1.0, "F01-02": .5, "F01-03": .5, "F01-04": .5, "F01-05": .5,
 "F02-01": .5, "F02-02": .5, "F03-01": .5, "F03-02": .5,
 "F04-01": .5, "F04-02": .5, "F04-03": .5,
 "F05-01": .5, "F05-02": .25, "F05-03": .25, "F05-04": .5, "F05-05": .5,
 "F06-01": .5, "F06-02": .25, "F06-03": .25,
 "F08-01": 1.0, "F08-02": .5, "F08-03": .5, "F08-04": .25, "F08-05": .25,
 "N1": .5, "N2": .25, "N3": 0, "N4": .5, "N5": .5,
}

# ---------- 功能名（读覆盖对照页） ----------
wst = wb["26项功能覆盖对照"]
FUNC_NAME, MOD_OF = {}, {}
for r in range(3, wst.max_row + 1):
    fid = wst.cell(row=r, column=1).value
    if fid and str(fid).startswith("F"):
        FUNC_NAME[str(fid)] = str(wst.cell(row=r, column=5).value)
        MOD_OF[str(fid)] = str(wst.cell(row=r, column=2).value)
# F04-04 异地灾备：Azure Blob 原生异地冗余（GRS/RA-GRS）自带，仅需配置与恢复演练（见WBS 5.2.8），不列入开发直估
assert set(FUNC_NAME) - {"F04-04"} == {f[0] for f in B} and "F04-04" in FUNC_NAME, "功能清单与覆盖页不一致"

# ---------- 渲染 ----------
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
F_TITLE = Font(bold=True, size=14, color="FFFFFF")
F_HDR = Font(bold=True, color="FFFFFF")
F_L1 = Font(bold=True)
FILL_NAVY = PatternFill("solid", fgColor="305496")
FILL_L1 = PatternFill("solid", fgColor="D9E1F2")
FILL_NFR = PatternFill("solid", fgColor="DDEBF7")
CENTER = Alignment(horizontal="center", vertical="center")

MOD_FILL = {}
_cur = None
_palette = ["EDEDF7", "E2EFDA", "FFF2CC", "FCE4D6", "F2F2F2", "E7E6E6", "E4DFEC"]
for fid in FUNC_NAME:
    if fid[:3] != _cur:
        _cur = fid[:3]
        MOD_FILL[_cur] = _palette[len(MOD_FILL) % len(_palette)]

if "功能拆解估算" in wb.sheetnames:
    del wb["功能拆解估算"]
wsn = wb.create_sheet("功能拆解估算", wb.sheetnames.index("功能视角分解") + 1)
wsn.sheet_properties.tabColor = "8FA9DB"
wsn.merge_cells("A1:L1")
wsn["A1"] = (f"功能拆解估算（子任务级直估）：模块→功能→子任务三层结构，{len(B)}项功能×{sum(len(s[1]) for s in B)}子任务"
             f" + 非功能与工程支撑×{sum(len(s[2]) for s in N)}子任务（F04-04异地灾备=Azure Blob原生能力，不计开发）— 独立估算，不锚定排期185人天")
wsn["A1"].font = F_TITLE; wsn["A1"].fill = FILL_NAVY; wsn["A1"].alignment = CENTER
wsn.row_dimensions[1].height = 26
GH = ["子任务编号", "模块", "功能/组名称", "子任务/工作包", "工作内容要点", "BA人天", "开发A·Wade", "开发B·DevB", "测试人天", "SLC人天（按功能）", "小计", "功能/组小计"]
GW = [11, 17, 19, 22, 40, 8, 10, 10, 8, 8, 8, 9]
for i, (h, w) in enumerate(zip(GH, GW), 1):
    wsn.column_dimensions[get_column_letter(i)].width = w
    c = wsn.cell(row=2, column=i, value=h); c.font = F_HDR; c.fill = FILL_NAVY; c.alignment = CENTER; c.border = BORDER

SECTIONS = [(fc, MOD_OF[fc], FUNC_NAME[fc], subs, PatternFill("solid", fgColor=MOD_FILL[fc[:3]])) for fc, subs in B] \
         + [(gc, "非功能与工程支撑", gn.replace("【非功能】", ""), subs, FILL_NFR) for gc, gn, subs in N]
n_subs = sum(len(s[3]) for s in SECTIONS)

r = 3
gt = [0.0] * 5
feat_tot = [0.0] * 5
nfr_tot = [0.0] * 5
sec_rows = []  # (模块名, r0, r1, fill)
for gcode, mname, gname, subs, fill in SECTIONS:
    is_nfr = gcode.startswith("N")
    r0 = r
    for n, (name, desc, ba, wa, db, te, sl) in enumerate(subs, 1):
        vals = [f"{gcode}-{n}", None, None, name, desc, ba or None, wa or None, db or None, te or None, None,
                f"=ROUND(SUM(F{r}:I{r}),2)"]
        for i, v in enumerate(vals, 1):
            c = wsn.cell(row=r, column=i, value=v); c.border = BORDER
            c.alignment = CENTER if i in (1, 6, 7, 8, 9, 10, 11) else Alignment(wrap_text=True, vertical="center")
            if i in (6, 7, 8, 9, 10, 11): c.number_format = "0.##"
        wsn.row_dimensions[r].height = 26
        r += 1
    r1 = r - 1
    # 功能/组名称（C列）与功能小计（L列）纵向合并
    wsn.merge_cells(start_row=r0, start_column=3, end_row=r1, end_column=3)
    cB = wsn.cell(row=r0, column=3)
    cB.value = gname
    cB.font = F_L1; cB.fill = fill
    cB.alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")
    wsn.merge_cells(start_row=r0, start_column=12, end_row=r1, end_column=12)
    gs = [round(sum(s[i] for s in subs), 2) for i in range(2, 6)]
    slc = SLC_BY_GRP.get(gcode, 0)
    cK = wsn.cell(row=r0, column=12, value=f"=ROUND(SUM(F{r0}:I{r1})+J{r0},2)")
    cK.font = F_L1; cK.fill = fill; cK.alignment = CENTER; cK.number_format = "0.##"
    wsn.merge_cells(start_row=r0, start_column=10, end_row=r1, end_column=10)
    cJ = wsn.cell(row=r0, column=10, value=slc or None)
    cJ.font = F_L1; cJ.fill = fill; cJ.alignment = CENTER; cJ.number_format = "0.##"
    for rr in range(r0, r1 + 1):
        for col in (2, 3, 10, 12):
            wsn.cell(row=rr, column=col).border = BORDER
    sec_rows.append((mname, r0, r1, fill))
    for i in range(4):
        gt[i] += gs[i]
        (nfr_tot if is_nfr else feat_tot)[i] += gs[i]
    gt[4] += slc
    (nfr_tot if is_nfr else feat_tot)[4] += slc

# 模块列（B列）：按模块跨功能合并
i = 0
while i < len(sec_rows):
    j = i
    while j + 1 < len(sec_rows) and sec_rows[j + 1][0] == sec_rows[i][0]:
        j += 1
    mname, r0, r1, fill = sec_rows[i][0], sec_rows[i][1], sec_rows[j][2], sec_rows[i][3]
    wsn.merge_cells(start_row=r0, start_column=2, end_row=r1, end_column=2)
    cM = wsn.cell(row=r0, column=2)
    cM.value = mname
    cM.font = F_L1; cM.fill = fill
    cM.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    i = j + 1

_last = r - 1
labels = ["", "总计", "", f"{len(B)}项功能×{sum(len(s[1]) for s in B)}子任务 + 非功能工程×{sum(len(s[2]) for s in N)}子任务（F04-04=Azure原生能力，不计开发）",
          "独立直估（另含少量需求/PM等工作未列入本表）",
          f"=ROUND(SUM(F3:F{_last}),2)", f"=ROUND(SUM(G3:G{_last}),2)", f"=ROUND(SUM(H3:H{_last}),2)",
          f"=ROUND(SUM(I3:I{_last}),2)", f"=ROUND(SUM(J3:J{_last}),2)", f"=ROUND(SUM(F{r}:J{r}),2)", ""]
for i, v in enumerate(labels, 1):
    c = wsn.cell(row=r, column=i, value=v); c.border = BORDER; c.fill = FILL_L1; c.font = F_L1
    c.alignment = CENTER if i not in (4, 5) else Alignment(horizontal="left", vertical="center")
    if i in (6, 7, 8, 9, 10, 11): c.number_format = "0.##"
wsn.freeze_panes = "D3"

# ---------- 对照说明 ----------
wsg = wb["功能视角分解"]
plan_feat = plan_common = 0.0
for rr in range(3, wsg.max_row + 1):
    a = str(wsg.cell(row=rr, column=1).value)
    v = wsg.cell(row=rr, column=14).value
    if a.startswith("F") and isinstance(v, (int, float)):
        plan_feat += v
    elif a.startswith("G") and isinstance(v, (int, float)):
        plan_common += v
grand = round(sum(gt), 2)
note = ("对照说明（四个口径）：① 功能直估 %s 人天（BA %s / Wade %s / DevB %s / 测试 %s / SLC %s）——已按配置界面/动态表单复杂度上调前端工作量；"
        "② 非功能与工程支撑直估 %s 人天（N1初始化与基座/N2 ESLint·SonarLint·SAST·依赖治理/N3性能·鉴权掩码·可观测·容量/N4部署运维）；"
        "③ 已识别原生组件能力（按配置+验证计价，子任务标注【原生】/【中间件】）：ADLS SSE/CMK加密、ADLS GRS异地冗余、Key Vault密钥管理、SQL DB PITR备份与Monitor告警、Entra ID+MSAL单点登录、App Insights/Log Analytics可观测性、Unity Catalog授权、Azure DevOps扫描门禁；"
        "④ 合计 %s 人天 vs 计划185：Wade %s（计划60，有余量）；DevB %s（超计划60约%s）；Mandy三职 %s（BA %s + 测试 %s + SLC %s，超计划50约%s）——角色级差异的吸收路径见下方「AI Native提效折算对照」。"
        % (round(sum(feat_tot), 2), feat_tot[0], feat_tot[1], feat_tot[2], feat_tot[3], feat_tot[4],
           round(sum(nfr_tot), 2), grand, gt[1], gt[2], round(gt[2] - 60, 2),
           round(gt[0] + gt[3] + gt[4], 2), gt[0], gt[3], gt[4], round(gt[0] + gt[3] + gt[4] - 50, 2)))
c = wsn.cell(row=r + 2, column=1, value=note)
c.font = Font(italic=True, size=9, color="C00000")
c.alignment = Alignment(wrap_text=True, vertical="top")
wsn.merge_cells(start_row=r + 2, start_column=1, end_row=r + 2, end_column=12)
wsn.row_dimensions[r + 2].height = 56

# AI Native提效折算对照（直估→计划的吸收路径）
mandy_g = round(gt[0] + gt[3] + gt[4], 2)
mandy_adj = round(gt[0] + gt[3] * 0.65 + gt[4] * 0.8, 2)
devb_adj = round(gt[2] * 0.85, 2)
_lines = [
    ("AI Native提效折算对照（直估→计划185的吸收路径）：", True),
    ("· Mandy：直估 %s（BA %s + 测试 %s + SLC %s，SLC已按功能级预估且开发侧文档移至N5）→ 测试×0.65（AI生成用例+人工审校执行）、SLC×0.8（合稿评审）、BA不折减 → ≈%s ≈ 计划50 ✓" % (mandy_g, gt[0], gt[3], gt[4], mandy_adj), False),
    ("· Wade：直估 %s（含N5开发侧文档2.5）≈ 计划60，基本持平 ✓" % (gt[1],), False),
    ("· DevB：直估 %s → ETL/SQL/作业配置AI生成×0.85 → ≈%s ≈ 计划60 ✓（最大风险假设：AI生成ETL需有效落地，纳入风险跟踪）" % (gt[2], devb_adj), False),
    ("· 合计：直估 %s → 折算 ≈%s + PM 15 = %s ≤ 185，余 ≈%s 人天缓冲；若提效不达标，启用备用赶工日（9/20、10/10）" % (grand, round(gt[1] + mandy_adj + devb_adj, 2), round(gt[1] + mandy_adj + devb_adj + 15, 2), round(185 - gt[1] - mandy_adj - devb_adj - 15, 2)), False),
]
rr2 = r + 4
for text, bold in _lines:
    c = wsn.cell(row=rr2, column=1, value=text)
    c.font = Font(bold=bold, size=9, color="1F4E79")
    c.alignment = Alignment(wrap_text=True, vertical="center")
    wsn.merge_cells(start_row=rr2, start_column=1, end_row=rr2, end_column=12)
    wsn.row_dimensions[rr2].height = 16
    rr2 += 1

# ---------- 使用说明（幂等） ----------
ws0 = wb["使用说明"]
desc = ("子任务级直估：模块→功能→子任务三层结构（模块列与功能列均为纵向合并单元格），26项功能×75子任务 + 非功能与工程支撑×%d子任务"
        "（N1框架初始化与基座/N2 ESLint·SonarLint·SAST代码质量治理/N3性能·鉴权·可观测·容量/N4部署运维），"
        "按BA/开发A/开发B/测试/SLC文档独立估人天；底部附与计划口径对照" % sum(len(s[2]) for s in N))
fv_seen = 0
for rr in range(1, ws0.max_row + 1):
    if ws0.cell(row=rr, column=1).value == "功能视角分解":
        fv_seen += 1
        if fv_seen > 1:
            ws0.cell(row=rr, column=1).value = None
            ws0.cell(row=rr, column=2).value = None
seen = 0
for rr in range(1, ws0.max_row + 1):
    if ws0.cell(row=rr, column=1).value == "功能拆解估算":
        if seen == 0:
            ws0.cell(row=rr, column=2).value = desc
            ws0.cell(row=rr, column=2).alignment = Alignment(wrap_text=True, vertical="top")
            ws0.row_dimensions[rr].height = 30
        else:
            ws0.cell(row=rr, column=1).value = None
            ws0.cell(row=rr, column=2).value = None
        seen += 1
if seen == 0:
    last = ws0.max_row + 1
    ws0.cell(row=last, column=1, value="功能拆解估算").font = Font(bold=True)
    v = ws0.cell(row=last, column=2, value=desc)
    v.alignment = Alignment(wrap_text=True, vertical="top")
    ws0.row_dimensions[last].height = 30

wb.save(PATH)
print(f"完成：{len(B)}功能×{sum(len(s[1]) for s in B)}子任务 + 非功能{len(N)}组×{sum(len(s[2]) for s in N)}子任务（共{n_subs}）")
print(f"功能直估: {round(sum(feat_tot),2)}（BA {feat_tot[0]} / Wade {feat_tot[1]} / DevB {feat_tot[2]} / 测试 {feat_tot[3]} / SLC {feat_tot[4]}）")
print(f"非功能直估: {round(sum(nfr_tot),2)}（BA {nfr_tot[0]} / Wade {nfr_tot[1]} / DevB {nfr_tot[2]} / 测试 {nfr_tot[3]} / SLC {nfr_tot[4]}）")
print(f"合计: {grand}（BA {gt[0]} / Wade {gt[1]} / DevB {gt[2]} / 测试 {gt[3]} / SLC {gt[4]}）vs 计划185")
