# -*- coding: utf-8 -*-
"""资源负荷与日历：人天矩阵改为公式（实时汇总自「WBS分解」），
改WBS工作量后本页自动重算。用法：python3 apply_formulas.py
"""
import openpyxl

PATH = "/home/user/wbs/RIMS项目WBS工作分解_AI Native版.xlsx"
wb = openpyxl.load_workbook(PATH)
ws = wb["WBS分解"]
last = max(r for r in range(4, ws.max_row + 1) if ws.cell(row=r, column=1).value)
S = "'WBS分解'"
rng = f"{S}!$C$4:$C${last}"
arn = f"{S}!$A$4:$A${last}"
ern = f"{S}!$E$4:$E${last}"
irn = f"{S}!$I$4:$I${last}"

ws3 = wb["资源负荷与日历"]
persons = {"Mark": 3, "Mandy": 4, "Wade": 5, "DevB": 6}  # 行号
for name, r in persons.items():
    for k in range(1, 9):  # 阶段1~8 -> B..I
        col = chr(ord("A") + k)
        f = (f'=IF(ROUND(SUMPRODUCT(({rng}=3)*({ern}="{name}")*(LEFT({arn},1)="{k}")*{irn}),2)=0,"",'
             f'ROUND(SUMPRODUCT(({rng}=3)*({ern}="{name}")*(LEFT({arn},1)="{k}")*{irn}),2))')
        ws3.cell(row=r, column=k + 1).value = f
    ws3.cell(row=r, column=10).value = f"=ROUND(SUM(B{r}:I{r}),2)"
for k in range(1, 9):
    col = chr(ord("A") + k)
    ws3.cell(row=7, column=k + 1).value = f"=ROUND(SUM({col}3:{col}6),2)"
ws3.cell(row=7, column=10).value = "=ROUND(SUM(J3:J6),2)"
wb.save(PATH)
print(f"资源负荷矩阵已公式化（引用WBS分解 4~{last} 行）")
