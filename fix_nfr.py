# -*- coding: utf-8 -*-
"""修复NFR链路两处时序不合理 + 全簿一致性更新"""
import datetime as dt
import openpyxl

PATH = "/home/user/wbs/RIMS项目WBS工作分解_AI Native版.xlsx"
wb = openpyxl.load_workbook(PATH)
ws = wb["WBS分解"]

# 行定位：A列编码 -> 行号
row_of = {}
for r in range(4, ws.max_row + 1):
    v = ws.cell(row=r, column=1).value
    if v: row_of[str(v)] = r

def setv(code, col, val):
    ws.cell(row=row_of[code], column=col, value=val)

D53, D54, D55, D56 = (dt.datetime(2026, 12, d) for d in (3, 4, 7, 8))

# ---- 修复1：NFR基线并入SRS（M1冻结含NFR）；2.3.1改为测试可执行化 ----
setv("2.2.1", 2, "AI生成SRS初稿（26项功能+非功能需求NFR：数据流图/归档范围矩阵/保留期矩阵+性能·安全·合规·可用性·容量指标）+人工修订")
setv("2.2.1", 4, "SRS V1.0（含NFR基线）")
setv("2.3.1", 2, "NFR测试可执行化（将SRS中NFR基线细化为可测验收标准：检索响应/导出吞吐/并发/越权与掩码旁路用例集/RPO·RTO/容量上限，供6.3.2/7.2.3/7.2.6对照）")
setv("2.3.1", 4, "NFR可测验收标准清单")
setv("2.3.1", 16, "NFR基线已随SRS于M1冻结；本任务只做测试化转译，不改变基线")
setv("3.2.8", 2, "F04与非功能设计定稿（加密/密钥/备份灾备方案+性能容量与「数据不出境」合规基线，对应2.2.1 NFR基线）")
setv("3.2.8", 11, "3.2.5;2.2.2")

# ---- 修复2：NFR专项验证提前到UAT之前；预演/预跑与UAT并行 ----
setv("7.2.2", 6, D55); setv("7.2.2", 7, D56); setv("7.2.2", 8, 2)
setv("7.2.2", 16, "12/7~12/8与UAT并行（staging环境，互不冲突）；回退预案验证")
setv("7.2.3", 6, D53); setv("7.2.3", 7, D54); setv("7.2.3", 8, 2)
setv("7.2.3", 11, "6.1.6")
setv("7.2.3", 16, "★先于UAT（12/7）完成：不达标则调优复测后再组织UAT；结果提交M6与生产就绪评审")
setv("7.2.5", 6, D55); setv("7.2.5", 7, D56); setv("7.2.5", 8, 2)
setv("7.2.5", 16, "12/7~12/8与UAT并行；全流程dry-run")
setv("7.2.6", 6, D53); setv("7.2.6", 7, D54); setv("7.2.6", 8, 2)
setv("7.2.6", 11, "6.2.5;7.2.4")
setv("7.2.6", 16, "★先于UAT（12/7）完成：越权/掩码旁路用例全跑+备份恢复与灾备切换演练；结果提交就绪评审")
setv("7.1.1", 2, "系统测试全量执行（26项逐项+汇总NFR专项结果（7.2.3/7.2.6），CC真实数据：检索/预览/导出/处置/审计）")

# ---- 里程碑页：M1/M6退出条件 ----
ws2 = wb["里程碑计划"]
for r in range(3, 10):
    if str(ws2.cell(row=r, column=1).value).startswith("M1"):
        ws2.cell(row=r, column=5, value="SRS评审通过、业务方签字（SRS含非功能需求NFR基线）；26项优先级锁定；追踪矩阵建立")
    if str(ws2.cell(row=r, column=1).value).startswith("M6"):
        ws2.cell(row=r, column=5, value="全量测试26项通过；性能专项（7.2.3）与安全可靠性专项（7.2.6）已于12/4前达标（对照NFR基线2.2.1/2.3.1）；业务方UAT签字（12/7）；缺陷关闭或书面接受；生产就绪评审通过")

# ---- WBS词典 2.3.1 条目 ----
ws4 = wb["WBS词典"]
for r in range(3, ws4.max_row + 1):
    if str(ws4.cell(row=r, column=1).value) == "2.3.1":
        ws4.cell(row=r, column=2, value="NFR测试可执行化")
        ws4.cell(row=r, column=3, value="将SRS中已冻结的NFR基线（性能/安全/合规/可用性/容量）细化为可测的验收标准与用例集，供集成初验（6.3.2）与专项验证（7.2.3/7.2.6）逐项对照。")
        ws4.cell(row=r, column=4, value="NFR可测验收标准清单")
        ws4.cell(row=r, column=5, value="每项NFR有量化指标、测试方法与通过阈值；与2.2.1基线一一对应，无新增无遗漏")

# ---- 使用说明 NFR链路文字 ----
ws0 = wb["使用说明"]
for r in range(1, ws0.max_row + 1):
    v = ws0.cell(row=r, column=2).value
    if v and "NFR基线（2.3.1" in str(v):
        ws0.cell(row=r, column=2, value="功能（26项）与非功能统一排程：NFR基线随SRS一并冻结（2.2.1，M1=9/24）→NFR测试可执行化（2.3.1）→设计期定方案（3.2.8）→集成期初验（6.3.2）→专项验证先于UAT完成（7.2.3性能/7.2.6安全可靠性，12/3~12/4）→生产就绪评审确认达标（7.2.7，12/9）")
        break

wb.save(PATH)

# ---- 复核：日期合法 + 每日负载 + 时序断言 ----
wb = openpyxl.load_workbook(PATH)
ws = wb["WBS分解"]
HOL = {dt.date(2026, 9, 25)} | {dt.date(2026, 10, d) for d in range(1, 8)}
WD = []
d = dt.date(2026, 9, 14)
while len(WD) < 62:
    if d.weekday() < 5 and d not in HOL: WD.append(d)
    d += dt.timedelta(days=1)
from collections import defaultdict
load = defaultdict(float)
for r in range(4, ws.max_row + 1):
    if ws.cell(row=r, column=3).value != 3: continue
    o = ws.cell(row=r, column=5).value
    if o in ("Mark", "Mandy") or str(ws.cell(row=r, column=1).value) in ("2.4", "1.2.5") or "缓冲" in str(ws.cell(row=r, column=2).value): continue
    s, e = ws.cell(row=r, column=6).value.date(), ws.cell(row=r, column=7).value.date()
    for x in (s, e):
        assert x in WD, f"非工作日 {x}"
    days = [x for x in WD if s <= x <= e]
    for x in days: load[(o, x)] += ws.cell(row=r, column=9).value / len(days)
over = [(k, round(v, 2)) for k, v in load.items() if v > 1.01]
assert not over, over
row_of = {str(ws.cell(row=r, column=1).value): r for r in range(4, ws.max_row + 1) if ws.cell(row=r, column=1).value}
def d_(c): return ws.cell(row=row_of[c], column=7).value.date()
assert d_("2.2.1") < d_("2.2.2") <= dt.date(2026, 9, 24), "NFR应随SRS在M1冻结"
assert d_("7.2.3") < dt.date(2026, 12, 7) and d_("7.2.6") < dt.date(2026, 12, 7), "NFR专项应先于UAT(12/7)完成"
assert d_("7.2.2") <= dt.date(2026, 12, 8) and d_("7.2.5") <= dt.date(2026, 12, 8)
print("修复完成并复核通过：NFR随M1冻结；7.2.3/7.2.6提前至12/3~12/4（UAT前）；7.2.2/7.2.5移至12/7~12/8（与UAT并行）；负载≤1；日期均合法")
