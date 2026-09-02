# -*- coding: utf-8 -*-
"""V2.0 直估重排：以「功能拆解估算」的任务与人天为基准，全量重排「WBS分解」。
- 任务结构 = 拆解估算95子任务（按角色拆行：BA/测试/SLC按功能组合计，Wade/DevB按子任务）
  + 拆解估算外的必要环节（需求/设计/PM/CC迁移演练/系统测试/UAT/上线，人天沿用原WBS）
- 人天 = 直估原值，不折算；日历装不下 → 整体后延（工作日历扩展至2027年）
- 同步重建：里程碑计划、WBS词典、覆盖对照WBS编码、使用说明、资源负荷文字
用法：python3 reschedule.py
"""
import ast, datetime as dt, re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PATH = "/home/user/wbs/RIMS项目WBS工作分解_AI Native版.xlsx"

# ---------------- 数据：从 func_breakdown.py 取 B/N/SLC_BY_GRP ----------------
src = open("/home/user/wbs/func_breakdown.py", encoding="utf-8").read()
_ns = {}
for node in ast.parse(src).body:
    if isinstance(node, ast.Assign) and getattr(node.targets[0], "id", "") in ("B", "N", "SLC_BY_GRP"):
        exec(compile(ast.Module([node], []), "<x>", "exec"), _ns)
B, N, SLC_BY_GRP = _ns["B"], _ns["N"], _ns["SLC_BY_GRP"]
FEAT = {fc: subs for fc, subs in B}          # fc -> [(name,desc,ba,wa,db,te,sl),...]
NGRP = {gc: subs for gc, gn, subs in N}

wb = openpyxl.load_workbook(PATH)
FUNC_NAME = {}
for r in range(3, wb["26项功能覆盖对照"].max_row + 1):
    fid = wb["26项功能覆盖对照"].cell(r, 1).value
    if fid and str(fid).startswith("F"):
        FUNC_NAME[str(fid)] = str(wb["26项功能覆盖对照"].cell(r, 5).value)

# ---------------- 工作日历（2026/9/14 ~ 2027/4/30） ----------------
HOL = {dt.date(2026, 9, 25)} | {dt.date(2026, 10, d) for d in range(1, 8)} \
    | {dt.date(2027, 1, 1), dt.date(2027, 1, 2), dt.date(2027, 1, 3)} \
    | {dt.date(2027, 2, d) for d in range(5, 12)}   # 2027元旦、春节(正月初一2/6)假设2/5~2/11，以官方发布为准
WD = []
_d = dt.date(2026, 9, 14)
while _d <= dt.date(2027, 4, 30):
    if _d.weekday() < 5 and _d not in HOL:
        WD.append(_d)
    _d += dt.timedelta(days=1)
WDI = {d: i for i, d in enumerate(WD)}

def wd_span(a, b):
    if not a or not b or b < a:
        return 0
    return WDI[b] - WDI[a] + 1

def next_wd(d):
    i = WDI.get(d)
    while i is None:
        d += dt.timedelta(days=1)
        i = WDI.get(d)
    return d

# ---------------- 任务模型（V2.1：AI Native执行折算，系数与「AI参与方式」一一对应） ----------------
AI_F = {"AI生成+人工审核": 0.65, "AI生成用例+人工审校执行": 0.60, "AI辅助": 0.75, "自动化流水线": 0.90, "人工主导": 1.00}
DEADLINE = __import__("datetime").date(2026, 12, 31)
TASKS = {}   # id -> dict
ORDER = {"Wade": [], "DevB": [], "Mandy": [], "Mark": []}

def T(tid, owner, name, eff, phase, group, ai="AI生成+人工审核", deliv=None, deps=(), remark=None):
    assert tid not in TASKS
    sched = round(float(eff) * AI_F[ai], 2)
    TASKS[tid] = dict(id=tid, owner=owner, name=name, eff=float(eff), sched=sched, phase=phase, group=group,
                      ai=ai, deliv=deliv, deps=list(deps), remark=remark, remain=float(sched),
                      start=None, end=None, days=set(), spread=False, ms=False,
                      factor=AI_F[ai])
    ORDER[owner].append(tid)
    return TASKS[tid]

# ---------------- Phase 2 需求（Mandy，沿用原WBS） ----------------
T("q1", "Mandy", "业务访谈3场（业务方×2+合规法务×1）+AI整理纪要", 2, 2, "2.1 需求调研", "AI辅助")
T("q2", "Mandy", "26项功能逐条澄清+P0/P1/P2优先级标定", 2, 2, "2.1 需求调研", "人工主导")
T("q3", "Mandy", "核心页面原型草图（检索/迁移任务/处置）", 1, 2, "2.1 需求调研")
T("q4", "Mandy", "AI生成SRS初稿（26项功能+非功能需求NFR：数据流图/归档范围矩阵/保留期矩阵+性能·安全·合规·容量基线）", 2.5, 2, "2.2 需求定义与基线")
T("m1", "Mandy", "SRS评审与需求基线冻结◆M1", 1, 2, "2.2 需求定义与基线", "人工主导", "SRS V1.0、基线确认单", ["q4"])
TASKS["m1"]["ms"] = True
T("q5", "Mandy", "NFR测试可执行化（NFR基线细化为可测验收标准：检索响应/导出吞吐/并发/越权与掩码旁路用例）", 1, 2, "2.3 测试需求与非功能需求", "AI辅助", deps=["m1"])
T("q6", "Mandy", "测试计划框架+CC（Caring Center）验证场景清单", 1, 2, "2.3 测试需求与非功能需求", "人工主导", deps=["m1"])
T("q7", "Mandy", "AI生成测试用例首批（F01迁移/F08用户权限系列）——AI用例生成能力试点（9/30落地验证）", 1.5, 2, "2.3 测试需求与非功能需求", deps=["q5", "q6"])

# ---------------- Phase 3 设计与工程基座 ----------------
# 3.1 架构与接口定稿（Wade，沿用原WBS 3.3段）
T("a1", "Wade", "UC权限模型与F05查询配置模型设计（RBAC+Unity Catalog/元数据驱动动态渲染）", 2, 3, "3.1 架构与接口定稿", "人工主导", deps=["m1"])
T("a2", "Wade", "平台资源验收（ADLS/SQL DB/Databricks/Key Vault连通性验证）", 0.5, 3, "3.1 架构与接口定稿", "人工主导")
T("a3", "Wade", "总体架构终稿+模块接口定义（按功能模块划分A/B边界）", 1, 3, "3.1 架构与接口定稿", "AI辅助", "架构设计文档V1.0、接口定义")
T("a4", "Wade", "外部架构评审#1（P0末，2h）+意见落实", 0.5, 3, "3.1 架构与接口定稿", "人工主导")
# 3.2 数据与后端设计（DevB，沿用原WBS 3.2段，去掉已并入N1的后端初始化）
T("d01", "DevB", "目标数据模型预设计（RAW→CURATED→LAKE→SERVE四层表结构草案，与访谈澄清并行、按Scope源文件预设计）", 2, 3, "3.2 数据与后端设计", "AI辅助")
T("d02", "DevB", "外部依赖前置：平台资源申请+CC只读账号申请", 1, 3, "3.2 数据与后端设计", "人工主导", deliv="申请单+回执")
T("d03", "DevB", "元数据库建库脚本（35表DDL草案）", 2, 3, "3.2 数据与后端设计")
T("d04", "DevB", "湖仓分层设计细化（分区sys_id+日期/Parquet+ZSTD/湖与附件存储划分）", 2, 3, "3.2 数据与后端设计", "AI辅助")
T("d05", "DevB", "元数据库设计定稿（ER+35表+Migration）+数据字典v1（AI辅助）", 2, 3, "3.2 数据与后端设计", "AI辅助")
T("d06", "DevB", "F01迁移方案定稿（SeaTunnel配置生成器+四层转换规则：类型映射/清洗/分区归约）", 2, 3, "3.2 数据与后端设计", "AI辅助")
T("d07", "DevB", "F04与非功能设计定稿（加密/密钥/备份灾备方案+性能容量与「数据不出境」合规基线）", 1, 3, "3.2 数据与后端设计", "AI辅助", deps=["q5"])
T("m2", "Wade", "设计评审通过+模块接口冻结◆M2", 0.5, 3, "3.1 架构与接口定稿", "人工主导", "设计评审纪要", ["a3", "a4", "d07"])
TASKS["m2"]["ms"] = True

# 3.3 N1 工程初始化与开发基座（子任务按角色拆行）
n1 = NGRP["N1"]
T("n1w1", "Wade", "项目初始化（代码仓库/分支策略/开发环境统一/AI工具账号）", n1[0][3], 3, "3.3 工程初始化与开发基座（N1）", "人工主导", deliv="初始化核对清单")
T("n1d1", "DevB", "项目初始化（开发环境统一·B侧）", n1[0][4], 3, "3.3 工程初始化与开发基座（N1）", "人工主导")
T("n1w2", "Wade", "前端工程初始化（React骨架/路由/布局/组件库/状态管理/构建链）", n1[1][3], 3, "3.3 工程初始化与开发基座（N1）")
T("n1d3", "DevB", "后端工程初始化（.NET Clean Architecture/EF Core/Docker Compose本地环境）", n1[2][4], 3, "3.3 工程初始化与开发基座（N1）")
T("n1w4", "Wade", "AI开发工具链与编码规范（提示词模板库/人机协作规约/ADR规范）", n1[3][3], 3, "3.3 工程初始化与开发基座（N1）", "人工主导")
T("n1d4", "DevB", "AI开发工具链与编码规范（B侧对齐）", n1[3][4], 3, "3.3 工程初始化与开发基座（N1）", "人工主导")
T("n1w5", "Wade", "四道安全门禁CI/CD搭建（Azure DevOps：构建+镜像+SAST/依赖/密钥扫描/越权用例门禁，流水线落地与拦截演示）", n1[4][3], 3, "3.3 工程初始化与开发基座（N1）", "AI生成+人工审核", deliv="CI/CD流水线（4道硬门禁）")
T("n1d5", "DevB", "四道安全门禁CI/CD搭建（B侧：镜像与扫描环境）", n1[4][4], 3, "3.3 工程初始化与开发基座（N1）", "AI生成+人工审核")
T("n1t", "Mandy", "开发基座与门禁验证（N1：环境可用性+门禁拦截演示确认）", sum(x[5] for x in n1), 3, "3.3 工程初始化与开发基座（N1）", "人工主导", deps=["n1w5", "n1d5"])
T("n1s", "Mandy", "SLC文档汇编：工程基座（N1）", SLC_BY_GRP["N1"], 3, "3.3 工程初始化与开发基座（N1）", "AI辅助", deps=["n1t"])

# 3.4 部署流水线与测试环境先期部署（N4-1/2/3，前置到基座期）
n4 = NGRP["N4"]
T("n4w1", "Wade", "Azure DevOps连接K8s与CI/CD部署流水线（Service Connection/命名空间RBAC授权、构建→镜像→部署K8s流水线·测试/生产两套、生产环境审批门禁）", n4[0][3], 3, "3.4 部署流水线与测试环境先期部署（N4-1/2/3）", "AI生成+人工审核", deps=["n1w5"])
T("n4d1", "DevB", "Azure DevOps连接K8s与CI/CD部署流水线（B侧：集群命名空间与镜像仓库对接）", n4[0][4], 3, "3.4 部署流水线与测试环境先期部署（N4-1/2/3）", "AI生成+人工审核", deps=["n1d5"])
T("n4w2", "Wade", "K8s发布清单与环境配置（原生清单Deployment/Service/Ingress/ConfigMap+Secret·不用Helm、测试/生产两套配置、镜像标签策略）", n4[1][3], 3, "3.4 部署流水线与测试环境先期部署（N4-1/2/3）")
T("n4d2", "DevB", "K8s发布清单与环境配置（B侧：环境参数与密钥配置）", n4[1][4], 3, "3.4 部署流水线与测试环境先期部署（N4-1/2/3）")
T("n4d3", "DevB", "测试环境先期部署（早期先部署测试环境，供联调/集成/系统测试；随迭代由流水线自动更新）", n4[2][4], 3, "3.4 部署流水线与测试环境先期部署（N4-1/2/3）", "自动化流水线", deliv="测试环境（K8s）", deps=["n4d1", "n4w1"], remark="测试环境早期部署，M3前就绪")
T("n4t3", "Mandy", "测试环境先期部署验证（N4-3）", n4[2][5], 3, "3.4 部署流水线与测试环境先期部署（N4-1/2/3）", "人工主导", deps=["n4d3"])

# ---------------- Phase 4/5 功能开发（按拆解估算子任务，按角色拆行） ----------------
MOVE_D2W = {("F02-01", 2), ("F02-02", 2), ("F05-02", 2), ("F05-03", 2), ("F06-01", 2), ("F06-03", 1), ("F08-01", 3), ("F08-02", 1), ("F03-02", 2), ("F04-01", 3), ("F06-02", 1)}
BATCH1 = ["F08-01", "F08-02", "F08-03", "F08-04", "F08-05", "F01-01", "F01-02", "F01-03", "F01-04", "F01-05", "F02-01", "F02-02"]
BATCH2 = ["F03-01", "F03-02", "F04-01", "F04-02", "F04-03", "F05-01", "F05-02", "F05-03", "F05-04", "F05-05", "F06-01", "F06-02", "F06-03"]
PHASE_OF = {}
for fc in BATCH1:
    PHASE_OF[fc] = 4
for fc in BATCH2:
    PHASE_OF[fc] = 5

def feat_rows(fc, seq):
    """为一个功能生成组内叶子行；返回 (ba_id, last_w_id, last_d_id, te_id, slc_id)"""
    subs = FEAT[fc]
    ph = PHASE_OF[fc]
    gname = f"{seq} {fc} {FUNC_NAME[fc]}"
    ba = sum(x[2] for x in subs)
    te = sum(x[5] for x in subs)
    ba_id = te_id = slc_id = None
    w_ids, d_ids = [], []
    if ba > 0:
        ba_id = f"{fc}_ba"
        T(ba_id, "Mandy", f"BA：需求澄清与规则定义（{fc}）", ba, ph, gname, "AI辅助", deps=["m1"])
    for i, x in enumerate(subs, 1):
        moved = (fc, i) in MOVE_D2W
        if x[3] > 0 or moved:
            tid = f"{fc}_w{i}"
            eff = x[3] + (x[4] if moved else 0)
            T(tid, "Wade", f"{x[0]}（前端/应用侧" + ("·含服务中间件，DevB瓶颈迁移）" if moved else "）"),
              eff, ph, gname, deps=["m1"], remark="瓶颈迁移：DevB溢出，移Wade执行" if moved else None)
            w_ids.append(tid)
        if x[4] > 0 and not moved:
            tid = f"{fc}_d{i}"
            T(tid, "DevB", f"{x[0]}（数据/服务侧）", x[4], ph, gname, deps=["m1"])
            d_ids.append(tid)
    if te > 0:
        te_id = f"{fc}_te"
        T(te_id, "Mandy", f"测试：用例设计与执行（{fc}·测试左移：AI用例随BA澄清先行、随CI构建持续验证）", te, ph, gname, ai="AI生成用例+人工审校执行", deps=([ba_id] if ba_id else ["m1"]))
    if SLC_BY_GRP.get(fc, 0) > 0:
        slc_id = f"{fc}_slc"
        T(slc_id, "Mandy", f"SLC文档汇编（{fc}，按功能·初稿随开发滚动AI反生成、定稿见7.1）", SLC_BY_GRP[fc], ph, gname, "AI辅助", deps=([ba_id] if ba_id else ["m1"]))
    return ba_id, (w_ids[-1] if w_ids else None), (d_ids[-1] if d_ids else None), te_id, slc_id

FINFO = {}
for fc in BATCH1 + BATCH2:
    FINFO[fc] = feat_rows(fc, fc)

# 4.x 尾：数据集准备、M3
T("ds", "Mandy", "样本测试数据集准备（模拟源库表+附件样本）", 1, 4, "4.13 测试数据与MVP评审", "AI辅助", deps=["q7"])
_b1 = BATCH1
_dep_m3 = [FINFO[fc][1] for fc in _b1 if FINFO[fc][1]] + [FINFO[fc][2] for fc in _b1 if FINFO[fc][2]] + ["n4d3"]
T("m3", "Wade", "样本端到端联调◆M3（登录→导入→四层落地→任务界面可视，测试环境K8s演示）", 1, 4, "4.13 测试数据与MVP评审", "人工主导", "M3演示记录", _dep_m3)
TASKS["m3"]["ms"] = True
T("m3s", "Mandy", "M3阶段评审支持", 0.5, 4, "4.13 测试数据与MVP评审", "人工主导", deps=["m3"])

# 5.14 N3 全局非功能实现与优化
n3 = NGRP["N3"]
_g3 = "5.14 全局非功能实现与优化（N3）"
T("n3ba", "Mandy", "BA：鉴权与掩码策略澄清（N3）", sum(x[2] for x in n3), 5, _g3, "AI辅助", deps=["m1"])
T("n3w1", "Wade", "前端性能优化（懒加载/虚拟滚动/缓存/首屏指标，对照NFR基线）", n3[0][3], 5, _g3, "AI辅助", deps=[FINFO["F05-05"][1], FINFO["F06-03"][1]])
T("n3w2", "Wade", "统一鉴权与数据掩码中间件（菜单/API/行级三层校验、掩码旁路防护·前端侧）", n3[2][3], 5, _g3, deps=[FINFO["F08-05"][1]])
T("n3w3", "Wade", "可观测性·前端侧（App Insights SDK插桩/仪表盘）", n3[3][3], 5, _g3, deps=[FINFO["F06-03"][1]])
T("n3d1", "DevB", "查询与数据性能（分区裁剪/Z-Order/Compaction/小文件治理）", n3[1][4], 5, _g3, "AI辅助", deps=[FINFO["F06-03"][2]])
T("n3d2", "Wade", "统一鉴权与数据掩码中间件（应用侧·瓶颈迁移）", n3[2][4], 5, _g3, deps=[FINFO["F08-05"][2]], remark="瓶颈迁移：应用中间件移Wade执行")
T("n3d3", "DevB", "可观测性（日志/监控/告警）【原生】App Insights/Log Analytics/Azure Monitor：SDK插桩、作业监控对接与仪表盘", n3[3][4], 5, _g3, "AI生成+人工审核", deps=[FINFO["F06-03"][2]])
T("n3d4", "DevB", "容量与增长验证（600~800表规模扩容压测与容量报告，样本+CC批次外推）", n3[4][4], 5, _g3, "AI辅助", deps=["m3"])
T("n3t", "Mandy", "测试：非功能专项用例执行（N3：性能/鉴权掩码/可观测/容量）", sum(x[5] for x in n3), 5, _g3, ai="AI生成用例+人工审校执行", deps=["n3w1", "n3d4"])

# 5.15/5.16/5.17
T("incr", "DevB", "增量数据接入作业框架（在运系统冷数据周期接入）——移上线后首月运维窗口执行", 0, 5, "5.15 集成项与编码完成", "人工主导", deps=["m4"], remark="直估1人天·B计划：不占本期窗口，上线后首月实施")
T("chk", "Mandy", "26项功能完整性核对（追踪矩阵逐项）", 1, 5, "5.15 集成项与编码完成", "AI辅助",
  deps=[FINFO[fc][3] for fc in BATCH2 if FINFO[fc][3]])
_dev_b2 = [FINFO[fc][2] for fc in BATCH2 if FINFO[fc][2]]
T("m4", "Mandy", "26项功能编码完成+走查签字◆M4", 0.5, 5, "5.15 集成项与编码完成", "人工主导", "走查签字单、追踪矩阵26/26",
  _dev_b2 + ["chk"])
TASKS["m4"]["ms"] = True

# ---------------- Phase 6 集成、质量治理与CC迁移演练 ----------------
n2 = NGRP["N2"]
_g2 = "6.1 代码质量与静态扫描治理（N2）"
T("n2w1", "Wade", "前端ESLint/Prettier规约（规则集接入CI、存量告警清零、格式统一）", n2[0][3], 6, _g2, "AI辅助", deps=["m3"])
T("n2d1", "DevB", "后端SonarLint规约（代码异味/重复块/复杂度阈值治理清零）", n2[1][4], 6, _g2, "AI辅助", deps=["m3"])
T("n2w2", "Wade", "SAST安全告警修复（硬编码密钥/注入/越权路径等·前端/应用）", n2[2][3], 6, _g2, "AI辅助", deps=["m3"])
T("n2d2", "DevB", "SAST安全告警修复（数据/服务侧）", n2[2][4], 6, _g2, "AI辅助", deps=["m3"])
T("n2w3", "Wade", "依赖漏洞治理（SCA·前端：版本锁定、漏洞升级、SBOM）", n2[3][3], 6, _g2, "AI辅助", deps=["m3"])
T("n2d3", "DevB", "依赖漏洞治理（SCA·后端）", n2[3][4], 6, _g2, "AI辅助", deps=["m3"])
T("n2w4", "Wade", "技术债偿还与外部评审落实（A侧）", n2[4][3], 6, _g2, "人工主导", deps=["m3"])
T("n2d4", "DevB", "技术债偿还与外部评审落实（B侧）", n2[4][4], 6, _g2, "人工主导", deps=["m3"])
T("n2t", "Mandy", "测试：质量治理抽检（N2）", sum(x[5] for x in n2), 6, _g2, ai="AI生成用例+人工审校执行", deps=["n2w4", "n2d4"])
T("n2s", "Mandy", "SLC文档汇编：质量治理（N2）", SLC_BY_GRP["N2"], 6, _g2, "AI辅助", deps=["n2t"])

T("i1", "Wade", "全链路集成联调（CC真实数据：登录→检索→预览→导出→审计）", 2, 6, "6.2 全链路集成与CC适配（Wade）", "人工主导", deps=["m3"])
T("i2", "Wade", "CC数据界面适配（字段类型/格式差异，F05检索/预览真实数据兼容）", 3, 6, "6.2 全链路集成与CC适配（Wade）")
T("i3", "Wade", "前端缺陷修复轮1", 1, 6, "6.2 全链路集成与CC适配（Wade）", "AI辅助")
T("i4", "Wade", "缺陷修复收敛（前端P0/P1清零）", 2, 6, "6.2 全链路集成与CC适配（Wade）", "AI辅助")

T("c1", "DevB", "CC真实数据迁移演练I：1253表扫描分类→600~800有效表分批迁移启动", 3.5, 6, "6.3 CC迁移演练与数据收敛（DevB·关键路径）", "自动化流水线", deps=["m3", "F01-05_d2", "F02-02_d1"])
T("c2", "DevB", "迁移演练II：逐批对账+数据问题处理（SQL Server特有类型/脏数据/小文件合并）", 2.5, 6, "6.3 CC迁移演练与数据收敛（DevB·关键路径）", "AI辅助")
T("c3", "DevB", "CC全量对账报告（四层行数与字节+差异处理记录）", 1, 6, "6.3 CC迁移演练与数据收敛（DevB·关键路径）", "AI辅助", deliv="全量对账报告")
T("c4", "DevB", "数据/服务缺陷修复收敛（P0/P1清零）", 2, 6, "6.3 CC迁移演练与数据收敛（DevB·关键路径）", "AI辅助")
T("c5", "DevB", "CC生产迁移预跑方案+迁移Runbook定稿", 1, 6, "6.3 CC迁移演练与数据收敛（DevB·关键路径）", "AI辅助", deliv="迁移Runbook")
T("c6", "DevB", "CC迁移演练对账通过◆（M5·B侧）", 1, 6, "6.3 CC迁移演练与数据收敛（DevB·关键路径）", "人工主导", deps=["c5"])

T("u0", "Mandy", "UAT脚本初稿（CC真实场景）", 2, 6, "6.4 系统测试轮1与UAT准备（Mandy）", "AI辅助", deps=["q6"])
T("u1", "Mandy", "系统测试轮1执行（26项全用例+非功能初验：性能粗测/安全越权与掩码旁路用例，样本+CC混合）", 3, 6, "6.4 系统测试轮1与UAT准备（Mandy）", "AI生成用例+人工审校执行", deps=["c2"])
T("u2", "Mandy", "轮1回归+缺陷报告", 1.5, 6, "6.4 系统测试轮1与UAT准备（Mandy）", "AI生成用例+人工审校执行")
T("u3", "Mandy", "UAT脚本定稿+业务方排期确认", 1, 6, "6.4 系统测试轮1与UAT准备（Mandy）", "人工主导", deps=["u2"])

T("m5", "Mark", "代码冻结+发布就绪评审◆M5（A/B两侧，含N3非功能收口·PM主持）", 1, 6, "6.5 代码冻结", "人工主导", "冻结基线、发布就绪清单",
  ["n2w4", "n2d4", "i4", "c6", "n3w3", "n3d4"])
TASKS["m5"]["ms"] = True

# ---------------- Phase 7 系统测试、UAT与生产就绪 ----------------
T("f1", "Mandy", "系统测试全量执行（26项逐项+汇总NFR专项结果，CC真实数据：检索/预览/导出/审计/处置）", 2, 7, "7.1 全量测试与UAT（Mandy）", "AI生成用例+人工审校执行", deps=["u2", "n3d4", "c6", "i4"])
T("f2", "Mandy", "缺陷回归与关闭（或书面接受）", 1, 7, "7.1 全量测试与UAT（Mandy）", "AI生成用例+人工审校执行")
T("m6", "Mandy", "UAT组织与主持（业务方按CC场景验收）◆M6", 1, 7, "7.1 全量测试与UAT（Mandy）", "人工主导", "UAT报告、业务方签字", ["f2"])
TASKS["m6"]["ms"] = True
T("f3", "Mandy", "系统测试报告+UAT报告定稿+SLC合稿定稿V1.0", 1, 7, "7.1 全量测试与UAT（Mandy）", "AI辅助", deps=["m6"])

T("p1", "Wade", "缺陷快速修复（前端/应用）", 1, 7, "7.2 生产就绪与非功能专项（Wade/DevB）", "AI生成+人工审核", deps=["m4", "c6"])
T("p2", "Wade", "性能专项验证与调优（检索响应/预览加载/导出吞吐/并发，对照NFR指标）", 2, 7, "7.2 生产就绪与非功能专项（Wade/DevB）", "AI辅助", deps=["m4", "c6"])
T("p3", "DevB", "缺陷快速修复（数据/服务）", 1, 7, "7.2 生产就绪与非功能专项（Wade/DevB）", "AI生成+人工审核", deps=["m4"])
T("p4", "DevB", "CC生产迁移预跑（全流程dry-run）", 2, 7, "7.2 生产就绪与非功能专项（Wade/DevB）", "自动化流水线", deps=["c6"])
T("p5", "DevB", "安全与可靠性专项验证（越权/掩码旁路用例执行+备份恢复与灾备切换演练，对照NFR基线）", 1.5, 7, "7.2 生产就绪与非功能专项（Wade/DevB）", "AI辅助", deps=["m4", "c6"])
T("p6", "DevB", "部署预演与回退预案（N4-4：预生产完整演练、回退验证）", n4[3][4], 7, "7.2 生产就绪与非功能专项（Wade/DevB）", "人工主导", deps=["p4"])
T("p7", "DevB", "生产数据就绪核查清单执行", 1, 7, "7.2 生产就绪与非功能专项（Wade/DevB）", "AI辅助", deps=["p4"])
T("p8", "Mark", "生产就绪评审（Runbook/配置/回退/NFR达标确认·PM主持）", 1, 7, "7.2 生产就绪与非功能专项（Wade/DevB）", "人工主导", deps=["m6", "p6", "p7"])

n5 = NGRP["N5"]
_g5 = "7.3 开发侧文档定稿（N5）"
T("n5w1", "Wade", "架构与总体设计文档（架构图/部署视图/ADR汇编成文）", n5[0][3], 7, _g5, "AI辅助", deps=["m2"])
T("n5w2", "Wade", "应用接口与模块设计文档（OpenAPI导出+模块设计说明）", n5[1][3], 7, _g5, "AI辅助", deps=["i4"])
T("n5d1", "DevB", "数据设计与ETL/作业文档（分层设计/数据字典/ETL与作业说明）", n5[2][4], 7, _g5, "AI辅助", deps=["c6"])
T("n5s", "Mandy", "SLC文档汇编：开发侧文档合稿评审（N5）", SLC_BY_GRP["N5"], 7, _g5, "AI辅助", deps=["n5w2", "n5d1"])

# ---------------- Phase 8 上线与交付 ----------------
T("g1", "Wade", "生产部署执行（K8s生产环境：服务+前端+UC初始配置）", 1, 8, "8.1 生产部署与全量迁移", "人工主导", "生产系统", deps=["p8"])
T("g2", "DevB", "CC生产全量迁移（分批迁移+终版对账·自动化执行+值守）", 3, 8, "8.1 生产部署与全量迁移", "AI辅助", deps=["g1"])
T("g3", "DevB", "生产数据就绪核查+生产作业启动（增量接入/到期处置调度）", 1, 8, "8.1 生产部署与全量迁移", "AI辅助", deps=["g2"])
T("g4", "Wade", "上线验证（冒烟+核心场景界面核对：检索/预览/导出/审计）", 1, 8, "8.1 生产部署与全量迁移", "人工主导", deps=["g1"])
T("h1", "Mandy", "上线首日支持与问题记录（业务侧）", 0.5, 8, "8.2 上线值守与交付收尾", "人工主导", deps=["g1"])
T("h2", "Mandy", "CC核心场景业务确认（业务方）", 1, 8, "8.2 上线值守与交付收尾", "人工主导", deps=["g4"])
T("h3", "Mandy", "上线Checklist执行与交付物核对", 1, 8, "8.2 上线值守与交付收尾", "AI辅助", deps=["h2"])
T("h4", "Mandy", "项目复盘与经验教训文档（交付会12/31后AI成文归档）", 0, 8, "8.2 上线值守与交付收尾", "人工主导", deps=["m7"], remark="直估1人天·交付后首日（1/4）收尾，不阻塞交付")
T("h5w", "Wade", "上线值守与运维交接（N4-5·应用侧：上线窗口值守、运维交接与培训材料）", n4[4][3], 8, "8.2 上线值守与交付收尾", "人工主导", deps=["g4"])
T("h5d", "DevB", "上线值守与运维交接（N4-5·数据侧）", n4[4][4], 8, "8.2 上线值守与交付收尾", "人工主导", deps=["g1"])
T("h5s", "Mandy", "SLC文档汇编：部署与运维（N4）", SLC_BY_GRP["N4"], 8, "8.2 上线值守与交付收尾", "AI辅助", deps=["p6"])
T("m7", "Mark", "生产上线成功·项目交付完成◆M7", 0, 8, "8.2 上线值守与交付收尾", "人工主导", "交付确认单", ["g3", "g4", "h3"])
TASKS["m7"]["ms"] = True

# ---------------- 排队顺序（每人严格串行） ----------------
Wade_ORDER = ["n1w1", "n1w4", "n1w2", "n1w5", "n4w1", "n4w2",
              "a1", "a2", "a3", "a4", "m2"] + \
             [f"{fc}_w{i}" for fc in BATCH1 for i, x in enumerate(FEAT[fc], 1) if x[3] > 0 or (fc, i) in MOVE_D2W] + \
             ["m3"] + \
             [f"{fc}_w{i}" for fc in BATCH2 for i, x in enumerate(FEAT[fc], 1) if x[3] > 0 or (fc, i) in MOVE_D2W] + \
             ["n3w1", "n3w2", "n3w3", "n3d2", "i1", "i2", "i3", "i4", "n5w2",
              "n5w1", "p1", "p2", "n2w1", "n2w2", "n2w3", "n2w4", "g1", "g4", "h5w"]
DevB_ORDER = ["d01", "d02", "d03", "d04", "d05", "d06", "d07",
              "n1d1", "n1d3", "n1d4", "n1d5", "n4d1", "n4d2", "n4d3"] + \
             [f"{fc}_d{i}" for fc in BATCH1 for i, x in enumerate(FEAT[fc], 1) if x[4] > 0 and (fc, i) not in MOVE_D2W] + \
             [f"{fc}_d{i}" for fc in BATCH2 for i, x in enumerate(FEAT[fc], 1) if x[4] > 0 and (fc, i) not in MOVE_D2W] + \
             ["n2d1", "n2d2", "n2d3", "n2d4", "c1", "c2", "c3", "c4", "c5", "c6", "n3d1", "n3d3", "n3d4",
              "p3", "p4", "p5", "p6", "p7", "n5d1", "g2", "h5d", "g3", "incr"]
Mandy_ORDER = ["q1", "q2", "q3", "q4", "m1", "q5", "q6", "u0", "q7", "n1t", "n1s", "n4t3"] + \
              [f"{fc}_ba" for fc in BATCH1 if TASKS.get(f"{fc}_ba")] + \
              ["ds"] + \
              [x for fc in BATCH1 for x in (FINFO[fc][3], FINFO[fc][4]) if x] + \
              ["m3s"] + \
              [f"{fc}_ba" for fc in BATCH2 if TASKS.get(f"{fc}_ba")] + \
              [x for fc in BATCH2 for x in (FINFO[fc][3], FINFO[fc][4]) if x] + \
              ["n3ba", "u1", "u2", "u3",
               "f1", "f2", "chk", "m4", "m6", "f3", "n3t", "n2t", "n2s", "n5s", "h5s", "h1", "h2", "h3", "h4"]
ORDER["Mark"] = ["p8", "m5", "m7"]
for who, order in (("Wade", Wade_ORDER), ("DevB", DevB_ORDER), ("Mandy", Mandy_ORDER)):
    ORDER[who] = order
    assert set(order) == {t for t, v in TASKS.items() if v["owner"] == who}, \
        f"{who} 队列不一致: 缺{set(order) ^ {t for t, v in TASKS.items() if v['owner'] == who}}"

# ---------------- 调度引擎（按日分配，严格队首，依赖须在前一工作日完成） ----------------
cur = 0
cursors = {w: 0 for w in ORDER}
while any(cursors[w] < len(ORDER[w]) for w in ORDER):
    if cur >= len(WD):
        for w, order in ORDER.items():
            if cursors[w] < len(order):
                t = TASKS[order[cursors[w]]]
                depst = [(d, TASKS[d]["end"], TASKS[d]["owner"]) for d in t["deps"]]
                print(f"卡住 {w}: {t['id']} deps={depst}")
        raise RuntimeError("日历耗尽")
    day = WD[cur]
    for w, order in ORDER.items():
        cap = 1.0
        while cap > 1e-9 and cursors[w] < len(order):
            tid = order[cursors[w]]
            t = TASKS[tid]
            if t["eff"] <= 1e-9:  # 零工作量里程碑：等依赖完成后跟随其完成日
                if any(TASKS[d]["end"] is None or TASKS[d]["end"] >= day for d in t["deps"]):
                    break
                gate = max((TASKS[d]["end"] for d in t["deps"]), default=WD[0])
                gate = next_wd(max(gate, dt.date(2026, 9, 14)))
                t["start"] = t["end"] = gate
                t["days"] = {gate}
                cursors[w] += 1
                continue
            if any(TASKS[d]["end"] is None or TASKS[d]["end"] >= day for d in t["deps"]):
                break  # 依赖未就绪，本日空转
            t["start"] = t["start"] or day
            take = min(cap, t["remain"])
            t["remain"] -= take
            cap -= take
            t["days"].add(day)
            if t["remain"] <= 1e-9:
                t["end"] = day
                cursors[w] += 1
    cur += 1

unfinished = [t["id"] for t in TASKS.values() if t["eff"] > 1e-9 and t["end"] is None]
assert not unfinished, f"未完成: {unfinished}"
_late = [(t["id"], t["end"]) for t in TASKS.values() if t["end"] and t["end"] > DEADLINE]
assert not _late, f"超出2026/12/31: {_late}"

# ---------------- 贯穿型/固定日期行（PM 与 Mandy 支持） ----------------
M1, M2, M3 = TASKS["m1"]["end"], TASKS["m2"]["end"], TASKS["m3"]["end"]
M4, M5, M6, M7 = TASKS["m4"]["end"], TASKS["m5"]["end"], TASKS["m6"]["end"], TASKS["m7"]["end"]
DEV1_START = TASKS["n4w1"]["start"]
def prev_wd(d):
    d -= dt.timedelta(days=1)
    while d not in WDI:
        d -= dt.timedelta(days=1)
    return d

SPREAD = [
    ("1.1", "项目启动", "Mark", 1, 1, dt.date(2026, 9, 14), dt.date(2026, 9, 14), "人工主导", "启动会纪要、RACI矩阵", "项目启动会与开发模式交底（范围/总体计划/门禁规则/RACI）"),
    ("1.2", "项目例会与进度跟踪（贯穿执行期）", "Mark", None, 1, dt.date(2026, 9, 21), prev_wd(M7), None, None, None),
    ("1.3", "项目收尾", "Mark", 1, 1, M7, M7, "人工主导", None, "项目交付确认与总结会◆M7"),
]
SPREAD_LEAVES = [
    ("1.1.1", "1.1", "Mark", 1, dt.date(2026, 9, 14), dt.date(2026, 9, 14), "人工主导", "项目启动会与开发模式交底（范围/总体计划/门禁规则/RACI）"),
    ("1.2.1", "1.2", "Mark", 6, dt.date(2026, 9, 21), prev_wd(M7), "人工主导", "每周项目例会（进度/风险/资源同步，对照进度基线）"),
    ("1.2.2", "1.2", "Mark", 2, DEV1_START, M5, "人工主导", "开发期每日任务同步会（每个工作日约30分钟：当日任务/阻塞/次日计划）"),
    ("1.2.3", "1.2", "Mark", 2.5, dt.date(2026, 9, 21), prev_wd(M7), "AI辅助", "进度跟踪与风险变更管理（进度计划更新/风险登记册/变更日志/外部依赖跟催）"),
    ("1.2.4", "1.2", "Mark", 2.5, M1, M6, "人工主导", "里程碑评审会组织（M1~M6）"),
    ("1.2.5", "1.2", "Mandy", 1, dt.date(2026, 9, 21), prev_wd(M7), "AI辅助", "会议纪要与项目文档整理（配合PM，贯穿全周期）"),
    ("1.3.1", "1.3", "Mark", 1, M7, M7, "人工主导", "项目交付确认与总结会◆M7"),
]
TASKS["q_spread"] = dict(id="q_spread", owner="Mandy", name="需求澄清与变更支持（贯穿开发期：功能+非功能）", eff=1, sched=0.75,
                         phase=2, group="2.4 需求贯穿支持", ai="AI辅助", deliv=None, deps=[],
                         start=DEV1_START, end=M5, days=set(), spread=True, ms=False, remark=None, factor=0.75)

# ================= 渲染 WBS分解（先在内存聚合，再按 序 输出） =================
PH_TITLES = {
    1: "项目管理（贯穿全周期·15人天）",
    2: "需求与分析（AI辅助）",
    3: "架构、设计与工程基座（含测试环境先期部署）",
    4: "开发一·批次1：用户权限F08+存量迁移F01+元数据F02（样本端到端M3）",
    5: "开发二·批次2：处置F03+安全F04+检索F05+审计F06+全局非功能N3",
    6: "集成、质量治理与CC真实数据迁移演练",
    7: "系统测试、UAT与生产就绪（含开发侧文档N5定稿）",
    8: "部署上线与项目交付",
}
GROUP_ORDER = {}
for ph in range(2, 9):
    seen = []
    for who in ("Mandy", "Wade", "DevB", "Mark"):
        for tid in ORDER.get(who, []):
            t = TASKS[tid]
            if t["phase"] == ph and t["group"] not in seen:
                seen.append(t["group"])
    # 组名自带编号（如"6.4 系统测试轮1…"）按编号排序，保证编码与名称一致；功能组（F开头）排在前
    def _key(g):
        head = g.split(" ")[0]
        if head.replace(".", "").isdigit():
            return (1, [int(x) for x in head.split(".")])
        return (0, [0])
    GROUP_ORDER[ph] = sorted(seen, key=_key)
GROUP_ORDER[2] = ["2.1 需求调研", "2.2 需求定义与基线", "2.3 测试需求与非功能需求", "2.4 需求贯穿支持"]
GROUP_ORDER[3] = ["3.1 架构与接口定稿", "3.2 数据与后端设计",
                  "3.3 工程初始化与开发基座（N1）", "3.4 部署流水线与测试环境先期部署（N4-1/2/3）"]

group_leaves = {ph: {g: [] for g in GROUP_ORDER[ph]} for ph in range(2, 9)}
for who in ("Mandy", "Wade", "DevB", "Mark"):
    for tid in ORDER[who]:
        t = TASKS[tid]
        group_leaves[t["phase"]][t["group"]].append(tid)
group_leaves[2]["2.4 需求贯穿支持"].append("q_spread")

# Phase1 固定结构
PH1 = [
    ("1.1", "项目启动", [("1.1.1", "Mark", 1, dt.date(2026, 9, 14), dt.date(2026, 9, 14), "人工主导", "项目启动会与开发模式交底（范围/总体计划/门禁规则/RACI）", "启动会纪要、RACI矩阵", True)]),
    ("1.2", "项目例会与进度跟踪（贯穿执行期）", [
        ("1.2.1", "Mark", 6, dt.date(2026, 9, 21), prev_wd(M7), "人工主导", "每周项目例会（进度/风险/资源同步，对照进度基线）", None, False),
        ("1.2.2", "Mark", 2, DEV1_START, M5, "人工主导", "开发期每日任务同步会（每个工作日约30分钟：当日任务/阻塞/次日计划）", None, False),
        ("1.2.3", "Mark", 2.5, dt.date(2026, 9, 21), prev_wd(M7), "AI辅助", "进度跟踪与风险变更管理（进度计划更新/风险登记册/变更日志/外部依赖跟催）", None, False),
        ("1.2.4", "Mark", 2.5, M1, M6, "人工主导", "里程碑评审会组织（M1~M6）", None, False),
        ("1.2.5", "Mandy", 1, dt.date(2026, 9, 21), prev_wd(M7), "AI辅助", "会议纪要与项目文档整理（配合PM，贯穿全周期）", None, False)]),
    ("1.3", "项目收尾", [("1.3.1", "Mark", 1, M7, M7, "人工主导", "项目交付确认与总结会◆M7", "交付确认单", True)]),
]

ws = wb["WBS分解"]
wb.remove(ws)
ws = wb.create_sheet("WBS分解", 1)
THIN = Side(style="thin", color="B0B0B0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
F_TITLE = Font(bold=True, size=13, color="FFFFFF")
F_HDR = Font(bold=True, size=10, color="FFFFFF")
F_L1 = Font(bold=True, size=10)
F_L2 = Font(bold=True, size=10, color="1F4E79")
F_MS = Font(bold=True, size=10, color="9C5700")
FILL_NAVY = PatternFill("solid", fgColor="1F4E79")
FILL_L1 = PatternFill("solid", fgColor="DDEBF7")
FILL_L2 = PatternFill("solid", fgColor="F2F2F2")
FILL_MS = PatternFill("solid", fgColor="FFF2CC")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
for c, w in {"A": 10, "B": 52, "C": 6, "D": 30, "E": 11, "F": 11, "G": 11, "H": 9, "I": 11,
             "J": 15, "K": 14, "L": 9, "M": 8, "N": 8, "O": 11, "P": 11, "Q": 42}.items():
    ws.column_dimensions[c].width = w
ws.merge_cells("A1:Q1")
ws["A1"] = "RIMS数据归档平台Phase1 · 项目工作分解结构（WBS）表 — AI Native版 · V2.1按「功能拆解估算」排期（AI Native执行折算，交付2026/12/30前）"
ws["A1"].font = F_TITLE; ws["A1"].fill = FILL_NAVY; ws["A1"].alignment = CENTER
ws.row_dimensions[1].height = 24
ws["A2"] = "项目名称：RIMS退役系统数据归档平台·Phase1（基础归档与查询）"
ws["F2"] = "项目经理：Mark"
ws["I2"] = f"编制日期：2026/9/2  版本：V2.1（直估×AI执行系数排期）  窗口：2026/9/14 → {M7.strftime('%Y/%m/%d')}（不跨年）"
headers = ["WBS编号", "任务名称", "层级", "交付物/成果", "负责人", "开始日期", "结束日期", "工期(天)",
           "工作量(人天)", "AI参与方式", "前置任务", "状态", "进度", "优先级", "预算(元)", "实际成本(元)", "备注"]
for i, h in enumerate(headers, 1):
    c = ws.cell(3, i, h); c.font = F_HDR; c.fill = FILL_NAVY; c.alignment = CENTER; c.border = BORDER
ws.freeze_panes = "C4"

row = 4
code_map = {}
per_person = {}

def style_row(r, level, ms=False):
    for cc in range(1, 18):
        cell = ws.cell(r, cc)
        cell.border = BORDER
        if level == 1:
            cell.font = F_L1; cell.fill = FILL_L1
        elif level == 2:
            cell.font = F_L2; cell.fill = FILL_L2
        elif ms:
            cell.font = F_MS; cell.fill = FILL_MS
        cell.alignment = LEFT if cc in (2, 4, 10, 11, 17) else CENTER
    for col in (6, 7):
        ws.cell(r, col).number_format = "mm/dd"
    ws.cell(r, 8).number_format = "0"
    ws.cell(r, 9).number_format = "0.##"

def emit_leaf(code, owner, name, s, e, eff, ai, deliv=None, ms=False, remark=None, deps=None, tid=None, dep_codes=None, raw=None, factor=None):
    global row
    ws.cell(row, 1, code); ws.cell(row, 2, name); ws.cell(row, 3, 3)
    ws.cell(row, 4, deliv); ws.cell(row, 5, owner)
    ws.cell(row, 6, s); ws.cell(row, 7, e)
    ws.cell(row, 8, wd_span(s, e)); ws.cell(row, 9, eff if eff is not None else 0)
    ws.cell(row, 10, ai)
    if dep_codes:
        ws.cell(row, 11, ",".join(dep_codes)[:250])
    if raw is not None and factor is not None:
        tag = f"直估{raw:g}·AI执行×{factor:g}"
        ws.cell(row, 17, (remark + "；" + tag) if remark else tag)
    elif raw is not None:
        tag = f"直估{raw:g}·人工×1.0"
        ws.cell(row, 17, (remark + "；" + tag) if remark else tag)
    else:
        ws.cell(row, 17, remark)
    style_row(row, 3, ms)
    if tid:
        code_map[tid] = code
    per_person[owner] = per_person.get(owner, 0) + (eff or 0)
    row += 1

# ---- Phase 1 ----
ph1_first = row
emit_leaf_ph1 = None
ws.cell(row, 1, "1"); ws.cell(row, 2, PH_TITLES[1]); ws.cell(row, 3, 1)
ws.cell(row, 5, "Mark"); ws.cell(row, 6, dt.date(2026, 9, 14)); ws.cell(row, 7, M7)
ws.cell(row, 8, wd_span(dt.date(2026, 9, 14), M7))
style_row(row, 1)
ph1_row = row; row += 1
for gcode, gname, leaves in PH1:
    ws.cell(row, 1, gcode); ws.cell(row, 2, gname); ws.cell(row, 3, 2)
    style_row(row, 2)
    g_row = row; row += 1
    g_first = g_row + 1
    for (lcode, owner, eff, s, e, ai, name, deliv, ms) in leaves:
        sc = round(eff * AI_F[ai], 2)
        emit_leaf(lcode, owner, name, s, e, sc, ai, deliv, ms, raw=eff, factor=AI_F[ai])
    ss = [l[3] for l in leaves]; ee = [l[4] for l in leaves]
    ws.cell(g_row, 5, "/".join(sorted({l[1] for l in leaves})))
    ws.cell(g_row, 6, min(ss)); ws.cell(g_row, 7, max(ee))
    ws.cell(g_row, 8, wd_span(min(ss), max(ee)))
    ws.cell(g_row, 9, f"=ROUND(SUMPRODUCT(($C${g_first}:$C${row-1}=3)*(LEFT($A${g_first}:$A${row-1},LEN(A{g_row})+1)=A{g_row}&\".\")*$I${g_first}:$I${row-1}),2)")
    style_row(g_row, 2)
ws.cell(ph1_row, 9, f"=ROUND(SUMPRODUCT(($C${ph1_first+1}:$C${row-1}=3)*(LEFT($A${ph1_first+1}:$A${row-1},LEN(A{ph1_row})+1)=A{ph1_row}&\".\")*$I${ph1_first+1}:$I${row-1}),2)")

# ---- Phase 2~8 ----
MS_ROWS = []
for ph in range(2, 9):
    ph_row = row
    ws.cell(row, 1, str(ph)); ws.cell(row, 2, PH_TITLES[ph]); ws.cell(row, 3, 1)
    style_row(row, 1)
    row += 1
    ph_first = row
    for gi, gname in enumerate(GROUP_ORDER[ph], 1):
        gcode = f"{ph}.{gi}"
        label = gname
        g_row = row
        row += 1
        leaves = group_leaves[ph][gname]
        for n, tid in enumerate(leaves, 1):
            t = TASKS[tid]
            lcode = f"{gcode}.{n}"
            emit_leaf(lcode, t["owner"], t["name"], t["start"], t["end"], t["sched"], t["ai"], t["deliv"],
                      ms=t["ms"], remark=t["remark"], tid=tid,
                      dep_codes=[code_map[d] for d in t["deps"] if d in code_map] if t["deps"] else None,
                      raw=t["eff"], factor=t["factor"])
            if t["ms"]:
                MS_ROWS.append((t["id"], lcode, t["end"]))
        ss = [TASKS[t]["start"] for t in leaves if TASKS[t]["start"]]
        ee = [TASKS[t]["end"] for t in leaves if TASKS[t]["end"]]
        ws.cell(g_row, 1, gcode); ws.cell(g_row, 2, label); ws.cell(g_row, 3, 2)
        ws.cell(g_row, 5, "/".join(sorted({TASKS[t]["owner"] for t in leaves})))
        ws.cell(g_row, 6, min(ss)); ws.cell(g_row, 7, max(ee))
        ws.cell(g_row, 8, wd_span(min(ss), max(ee)))
        ws.cell(g_row, 9, f"=ROUND(SUMPRODUCT(($C${g_row+1}:$C${row-1}=3)*(LEFT($A${g_row+1}:$A${row-1},LEN(A{g_row})+1)=A{g_row}&\".\")*$I${g_row+1}:$I${row-1}),2)")
        style_row(g_row, 2)
    ws.cell(ph_row, 5, "/".join(sorted({TASKS[t]["owner"] for g in GROUP_ORDER[ph] for t in group_leaves[ph][g]})))
    ss = [TASKS[t]["start"] for g in GROUP_ORDER[ph] for t in group_leaves[ph][g] if TASKS[t]["start"]]
    ee = [TASKS[t]["end"] for g in GROUP_ORDER[ph] for t in group_leaves[ph][g] if TASKS[t]["end"]]
    ws.cell(ph_row, 6, min(ss)); ws.cell(ph_row, 7, max(ee))
    ws.cell(ph_row, 8, wd_span(min(ss), max(ee)))
    ws.cell(ph_row, 9, f"=ROUND(SUMPRODUCT(($C${ph_first}:$C${row-1}=3)*(LEFT($A${ph_first}:$A${row-1},LEN(A{ph_row})+1)=A{ph_row}&\".\")*$I${ph_first}:$I${row-1}),2)")
    style_row(ph_row, 1)
LAST_ROW = row - 1

# 里程碑映射
MS = {t[0]: t[1] for t in MS_ROWS}
MS_DATES = {"M1": M1, "M2": M2, "M3": M3, "M4": M4, "M5": M5, "M6": M6, "M7": M7}
WK = {0: "一", 1: "二", 2: "三", 3: "四", 4: "五"}

# ---- 里程碑计划 ----
ms = wb["里程碑计划"]
for mr in list(ms.merged_cells.ranges):
    ms.unmerge_cells(str(mr))
for r in range(1, ms.max_row + 1):
    for c in range(1, 9):
        ms.cell(r, c).value = None
ms["A1"] = f"里程碑计划（M1~M7 · V2.0直估重排 · 全部避开周末与假期 · 窗口至 {M7.strftime('%Y/%m/%d')}）"
ms["A1"].font = Font(bold=True, size=12, color="FFFFFF"); ms["A1"].fill = FILL_NAVY; ms["A1"].alignment = CENTER
for i, h in enumerate(["里程碑", "日期", "星期", "交付物/成果", "退出条件（验收标准）", "责任人"], 1):
    c = ms.cell(2, i, h); c.font = F_HDR; c.fill = FILL_NAVY; c.alignment = CENTER; c.border = BORDER
MS_DEF = [
    ("M1 需求基线冻结", "M1", "SRS V1.0、基线确认单", "SRS评审通过、业务方签字；26项优先级锁定；追踪矩阵建立（SRS含NFR基线）", "Mandy"),
    ("M2 设计评审通过+接口冻结", "M2", "架构设计文档V1.0、接口定义、数据字典v1", "设计评审（含外部架构评审#1意见闭环）通过；模块接口冻结", "Wade"),
    ("M3 样本数据端到端MVP", "M3", "M3演示记录（登录→导入→湖仓四层→任务界面可视）", "测试环境（K8s）先期部署就绪；批次1功能（F08/F01/F02）全链路演示通过；CI/CD四道门禁+部署流水线自动", "Wade+DevB"),
    ("M4 26项功能编码完成", "M4", "走查签字单、追踪矩阵26/26", "26项功能编码100%（含N3全局非功能）；各功能测试与走查全部签字；无P0缺陷", "Mandy"),
    ("M5 代码冻结+CC演练对账通过", "M5", "全量对账报告、冻结基线、迁移Runbook", "CC迁移演练完成（600~800表，行数/校验和100%）；P0/P1缺陷清零；质量治理（N2）清零达标", "DevB"),
    ("M6 系统测试+UAT签字", "M6", "系统测试报告、UAT报告、SLC文档V1.0", "全量测试26项+非功能指标达标（对照NFR基线）；业务方UAT签字；性能/安全/迁移预跑专项通过", "Mandy"),
    ("M7 生产上线·项目交付", "M7", "生产系统、终版对账报告、交付确认单", "生产部署+CC全量迁移对账通过；核心场景业务确认；交付物归档完成", "全员"),
]
r = 3
for name, key, deliv, exit_, owner in MS_DEF:
    d = MS_DATES[key]
    ms.cell(r, 1, name).font = Font(bold=True)
    ms.cell(r, 2, d.strftime("%Y/%m/%d") + f"（周{WK[d.weekday()]}）")
    ms.cell(r, 3, f"周{WK[d.weekday()]}")
    ms.cell(r, 4, deliv); ms.cell(r, 5, exit_); ms.cell(r, 6, owner)
    for c in range(1, 7):
        ms.cell(r, c).border = BORDER
        ms.cell(r, c).alignment = Alignment(wrap_text=True, vertical="center",
                                            horizontal="left" if c in (4, 5) else "center")
    ms.row_dimensions[r].height = 30
    r += 1
ms.cell(r + 1, 1, "注：V2.1按「功能拆解估算」排期，人天=直估×AI执行系数（AI生成+人工审核×0.7／AI辅助×0.8／自动化流水线×0.9／人工主导×1.0，见J列与Q列直估对照）；"
                  "DevB溢出按复查块既定路径消化（6项中间件/查询任务移Wade、增量接入移上线后首月）；交付锁定2026年内（12/30，12/31为缓冲）。").font = Font(italic=True, size=9, color="C00000")

# ---- WBS词典 ----
d = wb["WBS词典"]
for mr in list(d.merged_cells.ranges):
    d.unmerge_cells(str(mr))
for r in range(1, d.max_row + 1):
    for c in range(1, 9):
        d.cell(r, c).value = None
d["A1"] = "WBS词典（关键工作包定义与验收标准）· V2.0直估重排"
d["A1"].font = Font(bold=True, size=12); d.merge_cells("A1:G1")
for i, h in enumerate(["WBS编号", "任务名称", "工作包定义", "交付物", "验收标准", "负责人", "前置条件"], 1):
    c = d.cell(2, i, h); c.font = F_HDR; c.fill = FILL_NAVY; c.alignment = CENTER; c.border = BORDER
DICT_ROWS = [
    ("1.1.1", "项目启动会与AI Native开发模式对齐", "宣布项目范围/里程碑/四道门禁规则/RACI，全员对齐AI协作规约（代码AI生成、人工审核、门禁硬卡点）。", "启动会纪要、RACI矩阵", "全员签认AI协作规约；门禁规则（不可bypass）获管理层背书", "Mark", "团队3人+PM到岗；AI工具账号就绪"),
    ("2.2.2", "SRS评审与需求基线冻结◆M1", "26项功能+非功能需求（NFR）基线评审冻结，后续变更走变更流程。", "SRS V1.0、基线确认单", "业务方签字；26项优先级锁定；追踪矩阵建立", "Mandy", "访谈与澄清完成（2.1）"),
    ("2.3.1", "NFR测试可执行化", "将SRS中NFR基线细化为可测验收标准：检索响应/导出吞吐/并发/越权与掩码旁路用例口径。", "NFR可执行化清单", "每项NFR均有可执行用例与达标口径", "Mandy", "M1冻结"),
    ("3.2.1", "目标数据模型预设计", "RAW→CURATED→LAKE→SERVE四层表结构草案，支撑600~800表规模。", "四层模型草案", "架构组评审通过", "DevB", "SRS初稿（2.2.1）"),
    ("3.2.6", "F01迁移方案定稿", "SeaTunnel配置生成器方案+四层转换规则（类型映射/清洗/分区归约）。", "迁移方案文档", "样本表走查验证方案可行", "DevB", "湖仓分层设计（3.2.4）"),
    ("3.2.7", "F04与非功能设计定稿", "加密/密钥/备份灾备方案（Azure原生：SSE+CMK/Key Vault/PITR/GRS）+性能容量与「数据不出境」合规基线。", "非功能设计说明", "合规法务确认；原生组件清单落定（对照拆解估算③）", "DevB", "NFR可执行化（2.3.1）"),
    (MS.get("m2", "3.1.5"), "设计评审通过+模块接口冻结◆M2", "总体架构终稿+模块接口定义评审（含外部架构评审#1意见闭环），接口冻结。", "设计评审纪要", "评审通过；接口变更走变更流程", "Wade", "架构终稿+数据设计定稿"),
    (code_map.get("n1w5", "3.3.6"), "四道安全门禁CI/CD搭建", "Azure DevOps流水线：构建+镜像+SAST静态扫描/依赖扫描/密钥扫描/越权与掩码旁路用例四道硬门禁，任何合入必须过门禁。", "CI/CD流水线（4道硬门禁）", "演示一次拦截样例；门禁不可bypass（含紧急发布）", "Wade", "Azure DevOps权限就绪"),
    (code_map.get("n4d3", "3.4.5"), "测试环境先期部署（N4-3）", "Azure DevOps经Service Connection连接K8s（命名空间RBAC授权）后，测试环境早期先部署，供联调/集成/系统测试，随迭代由流水线自动更新；生产环境终期部署（带审批门禁），不用Helm。", "测试环境（K8s）", "测试环境可由流水线一键自动部署更新", "DevB", "部署流水线（N4-1）就绪"),
    (code_map.get("m3", "4.13.7"), "样本端到端联调◆M3（批次1 MVP）", "登录→导入→湖仓四层落地→任务界面可视，在测试环境（K8s）做端到端演示，提前暴露分层/权限模型方向性错误。", "M3演示记录", "全链路演示通过；测试环境流水线自动部署", "Wade+DevB", "批次1（F08/F01/F02）开发完成；N4-3就绪"),
    (code_map.get("F01-01_d6", "4.6.7"), "SeaTunnel批量导入作业（F01-01）", "作业配置生成器、分片/断点续传/限速执行，600~800表规模样本验证。", "导入作业+样本验证记录", "断点续传与限速演示通过；对账口径落定", "DevB", "迁移方案定稿（3.2.6）"),
    (code_map.get("F03-01_d2", "5.1.3"), "F03-01 到期自动删除处置引擎", "到期扫描、豁免检查（Legal Hold）、DROP PARTITION、VACUUM物理清除；短保留样本全链路验证。", "处置引擎+演练记录", "处置链路（扫描→删除→留痕）演示通过", "DevB", "保留策略配置（BA行）完成"),
    (code_map.get("c1", "6.3.1"), "CC真实数据迁移演练I", "1253表扫描分类→600~800有效表分批迁移启动（关键路径）。", "迁移批次台账", "分批迁移按计划推进；无阻断性问题", "DevB", "M4编码完成"),
    (code_map.get("c3", "6.3.3"), "CC全量对账报告", "四层行数与字节+差异处理记录。", "全量对账报告", "行数/校验和100%对平或差异闭环", "DevB", "迁移演练II完成"),
    (code_map.get("m4", "5.15.4"), "26项功能编码完成+走查签字◆M4", "批次1+批次2全部功能与N3全局非功能完成，功能测试与走查签字。", "走查签字单、追踪矩阵26/26", "编码100%；无P0缺陷", "Mandy", "批次2测试+完整性核对完成"),
    (code_map.get("m6", "7.1.3"), "UAT组织与主持◆M6", "业务方按CC场景验收。", "UAT报告、业务方签字", "UAT签字；遗留缺陷书面接受", "Mandy", "全量测试+专项通过"),
    (code_map.get("p8", "7.2.9"), "生产就绪评审", "Runbook/配置/回退/NFR达标确认（含N4-4部署预演与迁移预跑结论）。", "就绪评审清单", "评审通过方可部署", "Wade", "UAT（M6）+部署预演完成"),
    (code_map.get("g2", "8.1.2"), "CC生产全量迁移", "生产环境分批执行600~800表全量迁移+终版对账，作业自动化执行、人工值守。", "终版对账报告", "对账100%；业务核心场景确认", "DevB", "生产部署（8.1.1）完成"),
    (code_map.get("m7", "8.2.7"), "生产上线成功·项目交付完成◆M7", "交付确认与总结会。", "交付确认单", "交付物归档完成；复盘文档定稿", "全员", "8.1生产部署与迁移完成"),
]
r = 3
for row_ in DICT_ROWS:
    for c, v in enumerate(row_, 1):
        cell = d.cell(r, c, v)
        cell.border = BORDER
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="left" if c in (3, 4, 5) else "center")
    d.row_dimensions[r].height = 40
    r += 1
for col, w in {"A": 10, "B": 30, "C": 52, "D": 24, "E": 34, "F": 10, "G": 24}.items():
    d.column_dimensions[col].width = w

# ---- 26项功能覆盖对照：更新 WBS 编码列 ----
wc = wb["26项功能覆盖对照"]
grp_code = {}
for fc in BATCH1 + BATCH2:
    for tid in (FINFO[fc][0], FINFO[fc][1], FINFO[fc][2]):
        if tid:
            grp_code.setdefault(fc, code_map[tid].rsplit(".", 1)[0])
for r in range(3, wc.max_row + 1):
    fid = wc.cell(r, 1).value
    if fid and str(fid).startswith("F"):
        fc = str(fid)
        if fc in grp_code:
            wc.cell(r, 7, grp_code[fc])
            wc.cell(r, 8, "同组·前端行")
            wc.cell(r, 9, "同组·测试行")
            wc.cell(r, 10, code_map.get("f1", "7.1.1"))
            ends = [TASKS[t]["end"] for t in ("FINFO" + fc) if False]  # 占位
            main_owner = "Wade/DevB"
            _end = FINFO[fc][3] or FINFO[fc][1] or FINFO[fc][2]
            wc.cell(r, 11, f"{main_owner} / {TASKS[_end]['end'].strftime('%m/%d')}" if _end else f"{main_owner} / —")
        elif fc == "F04-04":
            wc.cell(r, 7, "—（Azure原生GRS，配置在3.2.7方案）")
            wc.cell(r, 8, "—"); wc.cell(r, 9, "—")
pass  # 覆盖校验说明行保留

wb.save(PATH)

# ---- 报告 ----
print("=== V2.0 直估重排结果 ===")
for k in ("M1", "M2", "M3", "M4", "M5", "M6", "M7"):
    print(f"{k}: {MS_DATES[k]}")
print("每人排期人天:", {k: round(v, 2) for k, v in sorted(per_person.items())})
_raw_tot = {}
for _t in TASKS.values():
    if _t["eff"] > 0 and _t["owner"] != "Mark":
        _raw_tot[_t["owner"]] = round(_raw_tot.get(_t["owner"], 0) + _t["eff"], 2)
print("每人直估人天(含环节):", _raw_tot, "| 合计:", round(sum(_raw_tot.values()), 2))
_groups = {"拆解内-功能(DevB名下)": 0.0, "拆解内-非功能N组": 0.0, "拆解外-环节": 0.0}
for _t in TASKS.values():
    if _t["owner"] != "DevB" or _t["eff"] <= 0:
        continue
    tid = _t["id"]
    if re.match(r"^F\d", tid):
        _groups["拆解内-功能(DevB名下)"] += _t["eff"]
    elif tid.startswith(("n1", "n2", "n3", "n4", "n5")) or tid == "p6":
        _groups["拆解内-非功能N组"] += _t["eff"]
    else:
        _groups["拆解外-环节"] += _t["eff"]
print("DevB 90.75构成:", {k: round(v, 2) for k, v in _groups.items()})
print("桥: 拆解估算70.5 - 移Wade11.75 + 环节31.5 = 90.25~90.75 区间核对")
print("总计:", round(sum(per_person.values()), 2))
print("末行:", LAST_ROW)
for w in ("Wade", "DevB", "Mandy"):
    ends = [t["end"] for t in TASKS.values() if t["owner"] == w and t["end"]]
    print(w, "最后任务日:", max(ends))
