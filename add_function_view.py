# -*- coding: utf-8 -*-
"""在现有WBS工作簿上追加「功能视角分解」页（功能×角色矩阵，与185人天对账）
数据来源：工作簿自身「WBS分解」(任务/负责人/人天/日期) +「26项功能覆盖对照」(功能↔任务映射)
用法：python3 add_function_view.py
"""
import datetime as dt
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PATH = "/home/user/wbs/RIMS项目WBS工作分解_AI Native版.xlsx"
wb = openpyxl.load_workbook(PATH)

# ---------- 读取WBS分解 ----------
ws = wb["WBS分解"]
by_wbs = {}
lvl3_codes = []
for r in range(4, ws.max_row + 1):
    code = ws.cell(row=r, column=1).value
    if not code:
        continue
    lvl = ws.cell(row=r, column=3).value
    rec = dict(owner=ws.cell(row=r, column=5).value,
               s=ws.cell(row=r, column=6).value, e=ws.cell(row=r, column=7).value,
               eff=ws.cell(row=r, column=9).value, lvl=lvl)
    by_wbs[str(code)] = rec
    if lvl == 3:
        lvl3_codes.append(str(code))

# ---------- 读取26项功能覆盖对照 -> FMAP ----------
wst = wb["26项功能覆盖对照"]
FMAP = []
for r in range(3, wst.max_row + 1):
    fid = wst.cell(row=r, column=1).value
    if fid and str(fid).startswith("F"):
        FMAP.append((str(fid), str(wst.cell(row=r, column=5).value),
                     str(wst.cell(row=r, column=7).value),
                     str(wst.cell(row=r, column=8).value) or None,
                     str(wst.cell(row=r, column=9).value)))
assert len(FMAP) == 26

# ---------- 共享任务分摊组 ----------
SHARE = {
    "4.1.1": ["F08-01", "F08-02"], "4.1.2": ["F08-03", "F08-04"],
    "4.3.1": ["F01-01", "F01-02", "F01-03", "F01-04", "F01-05"],
    "4.3.2": ["F05-01", "F05-02", "F05-03"], "4.2.5": ["F01-04", "F01-05"],
    "4.2.6": ["F02-01", "F02-02"], "5.2.1": ["F02-01", "F02-02"],
    "5.1.3": ["F06-01", "F06-02", "F06-03"], "5.1.5": ["F03-01", "F03-02"],
    "5.1.7": ["F05-04", "F05-05"],
}
WT = {
 "F01-01": ("迁移范围/映射规则澄清与验收标准", "迁移任务管理界面（调度/进度）", "SeaTunnel分片/限速/断点批量导入", "导入用例+样本表验证", "用户手册F01章"),
 "F01-02": ("文件来源与离线介质场景澄清", "迁移任务界面（文件批次）", "多源文件/大文件/属性保留/包解析", "导入与中断恢复用例", "用户手册F01章"),
 "F01-03": ("源→标准字段映射规则定义", "规则配置界面", "规则配置服务（映射/转换/校验）", "规则执行用例", "用户手册F01章"),
 "F01-04": ("完整性基准口径定义", "校验结果展示", "迁移前基准记录（行数/校验和）", "基准记录用例", "用户手册F01章"),
 "F01-05": ("比对差异处理口径", "比对结果展示", "迁移后比对（SHA-256/差异报告）", "比对用例（CC对账预演）", "用户手册F01章"),
 "F02-01": ("元数据字段体系与必填规则", "—", "统一元数据模型API", "模块2走查", "用户手册F02章"),
 "F02-02": ("来源系统编码与绑定规则", "元数据管理界面（协作）", "来源系统信息绑定API", "模块2走查", "用户手册F02章"),
 "F03-01": ("保留期矩阵与到期处置策略", "保留策略配置界面", "处置引擎（扫描→DROP PARTITION→VACUUM）", "到期删除走查", "用户手册F03章"),
 "F03-02": ("Legal Hold审批流程定义", "Legal Hold冻结/解除界面", "Hold/Release+豁免+留痕", "冻结/解除走查", "用户手册F03章"),
 "F04-01": ("密级与加密合规要求", "—", "SSE+CMK+PII字段加密", "加密落盘验证", "用户手册F04章"),
 "F04-02": ("密钥保管与审计要求", "—", "Key Vault创建/更新/权限/审计", "密钥轮换用例", "用户手册F04章"),
 "F04-03": ("备份策略要求（RPO）", "—", "备份策略配置+副本创建", "备份恢复用例", "用户手册F04章"),
 "F04-04": ("灾备要求（RTO）", "—", "异地副本+恢复流程演练", "灾备切换用例", "用户手册F04章"),
 "F05-01": ("检索需求与查询配置模型", "检索页动态表单+关键词检索", "—", "检索用例", "用户手册F05章"),
 "F05-02": ("来源系统筛选需求", "按系统名称检索", "—", "筛选用例", "用户手册F05章"),
 "F05-03": ("时间维度检索需求", "按业务时间检索", "—", "时间范围用例", "用户手册F05章"),
 "F05-04": ("列表字段/排序/导出需求", "结果列表（分页/排序）+流式导出", "—", "大批量用例", "用户手册F05章"),
 "F05-05": ("预览格式与权限要求", "在线预览（SAS直连）+大文件优化", "—", "预览用例", "用户手册F05章"),
 "F06-01": ("审计合规要求（归档日志）", "归档操作日志+审计查询界面", "—", "审计走查", "用户手册F06章"),
 "F06-02": ("查询行为审计要求", "查询操作日志记录", "—", "审计走查", "用户手册F06章"),
 "F06-03": ("导出审计要求（范围/IP）", "导出操作日志记录", "—", "审计走查", "用户手册F06章"),
 "F08-01": ("角色权限矩阵（RBAC）定义", "角色权限配置（界面+API）", "—", "越权用例", "用户手册F08章"),
 "F08-02": ("账号生命周期策略", "用户账号管理（界面+API）", "—", "账号用例", "用户手册F08章"),
 "F08-03": ("SSO对接需求（Entra ID）", "SSO集成（登录页+后端链路）", "—", "SSO登录用例", "用户手册F08章"),
 "F08-04": ("身份映射规则", "身份自动识别与关联", "—", "身份关联用例", "用户手册F08章"),
 "F08-05": ("异常访问控制策略", "超时/锁定/失效控制", "—", "异常登录用例", "用户手册F08章"),
}
# 开发人天按功能归属任务分摊
dev_alloc = {f[0]: {"Wade": 0.0, "DevB": 0.0} for f in FMAP}
for fcode, fname, main_w, coop_w, walk_w in FMAP:
    for code in [main_w] + ([coop_w] if coop_w and coop_w != "—" else []):
        t = by_wbs[code]
        members = SHARE.get(code, [fcode])
        dev_alloc[fcode][t["owner"]] = round(dev_alloc[fcode][t["owner"]] + t["eff"] / len(members), 4)
for code, members in SHARE.items():
    tot = by_wbs[code]["eff"]
    s = sum(round(tot / len(members), 4) for _ in members)
    dev_alloc[members[-1]][by_wbs[code]["owner"]] = round(dev_alloc[members[-1]][by_wbs[code]["owner"]] + tot - s, 4)
# SHARE中未被FMAP主/协作列直接引用的任务（如5.1.7导出优化服务F05-04/05），按成员均摊补记
_covered = {f[2] for f in FMAP} | {f[3] for f in FMAP if f[3] and f[3] != "—"}
EXTRA_WBS = {}
for code, members in SHARE.items():
    if code not in _covered:
        t = by_wbs[code]
        for fid in members:
            dev_alloc[fid][t["owner"]] = round(dev_alloc[fid][t["owner"]] + t["eff"] / len(members), 4)
            EXTRA_WBS.setdefault(fid, []).append(code)

# Mandy模块池（对应真实任务），模块内按0.25步长顺序分配
MOD_FUNCS = {}
for f in FMAP:
    MOD_FUNCS.setdefault(f[0].split("-")[0], []).append(f[0])
def spread(amt, fids, step=0.25):
    out, rem = {f: 0.0 for f in fids}, amt
    for f in fids:
        take = min(step, rem); out[f] = round(take, 2); rem = round(rem - take, 2)
        if rem <= 0: break
    if rem > 0: out[fids[-1]] = round(out[fids[-1]] + rem, 2)
    return out
BA_POOL = {"2.1.2": {"F01": .5, "F02": .25, "F03": .25, "F04": .25, "F05": .25, "F06": .25, "F08": .25}}
TEST_POOL = {"2.3.3": {"F01": .75, "F08": .75}, "4.4.1": {"F05": .75, "F06": .75},
             "4.4.3": {"F02": .25, "F03": .25, "F04": .25, "F05": .25}, "4.4.4": {"F01": .5, "F08": .5},
             "5.3.2": {"F02": 1}, "5.3.4": {"F05": 1}, "5.3.6": {"F06": 1},
             "5.3.9": {"F03": .5, "F04": .5}, "5.3.5": {"F05": .5, "F06": .5}}
DOC_POOL = {"4.4.5": {"F01": .75, "F08": .75}, "5.3.1": {"F01": .75, "F08": .75}, "5.3.3": {"F02": .25, "F03": .25}}
for pool in (BA_POOL, TEST_POOL, DOC_POOL):
    for code, dist in pool.items():
        assert round(sum(dist.values()), 2) == by_wbs[code]["eff"], f"池校验失败 {code}"
m_alloc = {f[0]: {"BA": 0.0, "测试": 0.0, "文档": 0.0} for f in FMAP}
for pool, col in ((BA_POOL, "BA"), (TEST_POOL, "测试"), (DOC_POOL, "文档")):
    for code, dist in pool.items():
        for mod, amt in dist.items():
            for fid, v in spread(amt, MOD_FUNCS[mod]).items():
                m_alloc[fid][col] = round(m_alloc[fid][col] + v, 2)

PUBLIC = [
 ("G1", "需求与SRS（跨功能）", {"BA": ["2.1.1", "2.1.3", "2.2.1", "2.2.2", "2.3.1", "2.4.1"]},
  {"BA": "业务访谈3场/原型草图/SRS生成与评审/NFR基线/需求贯穿"}),
 ("G2", "项目初始化与工程基座", {"Wade": ["3.1.1", "3.1.2", "3.1.3", "3.1.4"], "DevB": ["3.2.1"]},
  {"Wade": "仓库/分支/环境/AI规范/四道门禁CI-CD/前端骨架", "DevB": "后端骨架+本地Docker环境"}),
 ("G3", "架构与非功能设计", {"Wade": ["3.3.2", "3.3.3", "3.3.4", "3.3.5", "3.3.6"],
   "DevB": ["3.2.2", "3.2.3", "3.2.4", "3.2.5", "3.2.6", "3.2.7", "3.2.8"]},
  {"Wade": "UC权限与查询模型/资源验收/架构终稿/外部评审#1/M2", "DevB": "四层表结构/DDL/分层设计/迁移方案/安全与非功能设计"}),
 ("G4", "数据与测试基础", {"Wade": ["5.1.6"], "DevB": ["4.2.4", "4.2.7", "5.2.9"], "测试": ["2.3.2", "4.4.2"], "文档": ["5.3.7", "5.3.10"]},
  {"Wade": "技术债偿还（A侧）", "DevB": "RAW/四层落地/SERVE物化/增量接入框架", "测试": "测试计划/样本数据集", "文档": "手册合稿/操作手册"}),
 ("G5", "系统集成与CC真实数据迁移", {"Wade": ["4.3.3", "5.1.8", "5.1.9", "6.1.1", "6.1.2", "6.1.3", "6.1.4", "6.1.5", "6.1.6"],
   "DevB": ["5.2.10", "5.2.11", "6.2.1", "6.2.2", "6.2.3", "6.2.4", "6.2.5", "6.2.6"], "测试": ["6.3.2", "6.3.3"]},
  {"Wade": "M3端到端联调/全链路走查/CC界面适配/缺陷收敛/发布包/代码冻结M5", "DevB": "CC迁移演练/逐批对账/全量对账报告/预跑", "测试": "系统测试轮1+回归"}),
 ("G6", "系统测试与UAT", {"Wade": ["7.2.1"], "DevB": ["7.2.4"],
   "测试": ["4.4.6", "5.3.8", "5.3.11", "5.3.12", "5.3.13", "5.4.2", "6.3.1", "6.3.4", "7.1.1", "7.1.2", "7.1.3"],
   "文档": ["6.3.5", "7.1.4", "8.2.4"]},
  {"Wade": "前端/应用缺陷修复", "DevB": "数据/服务缺陷修复", "测试": "UAT脚本/全量测试/UAT组织/26项核对", "文档": "手册评审定稿/复盘"}),
 ("G7", "非功能专项验证", {"Wade": ["5.1.4", "5.4.1", "7.2.2", "7.2.3", "7.2.7"], "DevB": ["5.2.4", "7.2.5", "7.2.6", "7.2.8"]},
  {"Wade": "性能粗调/部署预演/性能专项/就绪评审/外部评审#2", "DevB": "ADLS分层/迁移预跑/安全可靠性专项/数据核查"}),
 ("G8", "部署上线与值守", {"Wade": ["8.1.1", "8.1.4", "8.2.5"], "DevB": ["8.1.2", "8.1.3", "8.2.6"], "测试": ["8.2.1", "8.2.2"], "文档": ["8.2.3"]},
  {"Wade": "生产部署/冒烟验证/应用侧值守", "DevB": "CC全量迁移/作业启动/数据侧值守", "测试": "上线支持/业务确认", "文档": "交付物核对"}),
 ("G9", "项目管理（Mark）", {"PM": ["1.1.1", "1.2.1", "1.2.2", "1.2.3", "1.2.4", "1.3.1", "8.2.7"], "文档": ["1.2.5"]},
  {"PM": "启动会/周例会/每日同步/进度风险跟踪/里程碑评审/交付确认", "文档": "会议纪要与项目文档整理（配合PM）"}),
]
# 归属完整性：每个三级任务恰在一处
_attr = set()
for f in FMAP:
    _attr.update([f[2]] + ([f[3]] if f[3] and f[3] != "—" else []))
_attr |= set(SHARE)
for pool in (BA_POOL, TEST_POOL, DOC_POOL):
    _attr |= set(pool)
_pub = {c for g in PUBLIC for col, codes in g[2].items() for c in codes}
assert not _attr & _pub, f"双重归属: {_attr & _pub}"
_miss = [c for c in lvl3_codes if c not in _attr and c not in _pub]
assert not _miss, f"未归属: {_miss}"

# ---------- 样式 ----------
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
NAVY = "305496"; L1FILL = "D9E1F2"; GRAY = "F2F2F2"
F_TITLE = Font(bold=True, size=14, color="FFFFFF")
F_HDR = Font(bold=True, color="FFFFFF")
F_L1 = Font(bold=True)
FILL_NAVY = PatternFill("solid", fgColor=NAVY)
FILL_L1 = PatternFill("solid", fgColor=L1FILL)
FILL_G = PatternFill("solid", fgColor=GRAY)
CENTER = Alignment(horizontal="center", vertical="center")

if "功能视角分解" in wb.sheetnames:
    del wb["功能视角分解"]
wsg = wb.create_sheet("功能视角分解", wb.sheetnames.index("26项功能覆盖对照") + 1)
wsg.sheet_properties.tabColor = "C00000"
wsg.merge_cells("A1:P1")
wsg["A1"] = "功能视角分解（功能×角色矩阵）：每个功能由谁做什么、投入多少人天 — 与WBS/185人天完全对账"
wsg["A1"].font = F_TITLE; wsg["A1"].fill = FILL_NAVY; wsg["A1"].alignment = CENTER
wsg.row_dimensions[1].height = 26
GH = ["功能编号", "模块", "具体功能（与源文件一致）",
      "BA·需求（Mandy）", "BA人天", "开发A·Wade", "A人天", "开发B·DevB", "B人天",
      "测试·验证（Mandy）", "测试人天", "文档（Mandy）", "文档人天", "小计人天", "关联WBS任务", "计划完成"]
GW = [8, 13, 20, 24, 7, 24, 7, 24, 7, 22, 7, 14, 7, 8, 16, 10]
for i, (h, w) in enumerate(zip(GH, GW), 1):
    wsg.column_dimensions[get_column_letter(i)].width = w
    c = wsg.cell(row=2, column=i, value=h); c.font = F_HDR; c.fill = FILL_NAVY; c.alignment = CENTER; c.border = BORDER
MOD_NAME = {"F01": "模块1·存量迁移", "F02": "模块2·元数据", "F03": "模块3·生命周期处置", "F04": "模块4·存储安全",
            "F05": "模块5·检索利用", "F06": "模块6·审计合规", "F08": "模块8·全局治理"}
_mod_fills = {"F01": "EDEDF7", "F02": "E2EFDA", "F03": "FFF2CC", "F04": "FCE4D6", "F05": "DDEBF7", "F06": "E7E6E6", "F08": "E4DFEC"}
rg = 3
tot_cols = {"BA": 0.0, "Wade": 0.0, "DevB": 0.0, "测试": 0.0, "文档": 0.0, "PM": 0.0}
def _d(x):
    return x.date() if isinstance(x, dt.datetime) else x
for fcode, fname, main_w, coop_w, walk_w in FMAP:
    coop = coop_w if coop_w and coop_w != "—" else None
    ba, wa, db, te, do = WT[fcode]
    ev = {k: round(v, 2) for k, v in {**m_alloc[fcode], **dev_alloc[fcode]}.items()}
    vals = [fcode, MOD_NAME[fcode[:3]], fname,
            ba, ev["BA"] or None, wa, ev["Wade"] or None, db, ev["DevB"] or None,
            te, ev["测试"] or None, do, ev["文档"] or None,
            round(sum(ev.values()), 2),
            main_w + (";" + coop if coop else "") + (";" + ";".join(EXTRA_WBS.get(fcode, [])) if EXTRA_WBS.get(fcode) else ""),
            max(_d(by_wbs[c]["e"]) for c in [main_w, coop, walk_w] + EXTRA_WBS.get(fcode, []) if c).strftime("%m/%d")]
    fill = PatternFill("solid", fgColor=_mod_fills[fcode[:3]])
    for i, v in enumerate(vals, 1):
        c = wsg.cell(row=rg, column=i, value=v); c.border = BORDER
        c.alignment = CENTER if i in (1, 5, 7, 9, 11, 13, 14, 16) else Alignment(wrap_text=True, vertical="center")
        c.fill = fill
        if i in (5, 7, 9, 11, 13, 14): c.number_format = "0.##"
    for k in ("BA", "Wade", "DevB", "测试", "文档"):
        tot_cols[k] += ev[k]
    wsg.row_dimensions[rg].height = 34
    rg += 1
for gid, gname, cols, desc in PUBLIC:
    sums = {k: round(sum(by_wbs[c]["eff"] for c in v), 2) for k, v in cols.items()}
    all_codes = [c for v in cols.values() for c in v]
    vals = [gid, "公共支撑", gname,
            desc.get("BA", "—"), sums.get("BA") or None, desc.get("Wade", "—"), sums.get("Wade") or None,
            desc.get("DevB", "—"), sums.get("DevB") or None, desc.get("测试", "—"), sums.get("测试") or None,
            desc.get("文档", "—"), sums.get("文档") or None,
            round(sum(sums.values()), 2), ",".join(all_codes[:4]) + ("…" if len(all_codes) > 4 else ""),
            max(_d(by_wbs[c]["e"]) for c in all_codes).strftime("%m/%d")]
    for i, v in enumerate(vals, 1):
        c = wsg.cell(row=rg, column=i, value=v); c.border = BORDER; c.fill = FILL_G
        c.alignment = CENTER if i in (1, 5, 7, 9, 11, 13, 14, 16) else Alignment(wrap_text=True, vertical="center")
        if i in (5, 7, 9, 11, 13, 14): c.number_format = "0.##"
    for k in ("BA", "Wade", "DevB", "测试", "文档", "PM"):
        tot_cols[k] += sums.get(k, 0)
    wsg.row_dimensions[rg].height = 34
    rg += 1
grand = round(sum(tot_cols[k] for k in ("BA", "Wade", "DevB", "测试", "文档")) + tot_cols["PM"], 2)
mandy_tot = round(tot_cols["BA"] + tot_cols["测试"] + tot_cols["文档"], 2)
assert grand == 185 and tot_cols["Wade"] == 60 and tot_cols["DevB"] == 60, f"对账失败 {tot_cols} grand={grand}"
assert mandy_tot == 50, f"Mandy对账失败 {mandy_tot}"
labels = ["合计", "", "26项功能 + 9类公共支撑",
          "Mandy·BA", tot_cols["BA"], "Wade（开发A）", tot_cols["Wade"], "DevB（开发B）", tot_cols["DevB"],
          "Mandy·测试", tot_cols["测试"], "Mandy·文档", tot_cols["文档"], grand, "", "PM Mark " + str(tot_cols["PM"])]
for i, v in enumerate(labels, 1):
    c = wsg.cell(row=rg, column=i, value=v); c.border = BORDER; c.fill = FILL_L1; c.font = F_L1
    c.alignment = CENTER
    if i in (5, 7, 9, 11, 13, 14): c.number_format = "0.##"
wsg.freeze_panes = "D3"
wsg.auto_filter.ref = "A2:P" + str(rg)
note = wsg.cell(row=rg + 1, column=1,
    value="对账：Mandy 50（BA %s + 测试 %s + 文档 %s）+ Wade 60 + DevB 60 + PM 15 = 185人天，与「WBS分解」「资源负荷与日历」完全一致；开发人天按关联WBS任务实际工作量分摊（共享任务按功能均分）；人天为空=该角色此功能无直接投入（工作量计入公共支撑行）"
          % (tot_cols["BA"], tot_cols["测试"], tot_cols["文档"]))
note.font = Font(italic=True, size=9, color="C00000")

# 使用说明追加一行
ws0 = wb["使用说明"]
last = ws0.max_row + 1
k = ws0.cell(row=last, column=1, value="功能视角分解")
k.font = Font(bold=True)
v = ws0.cell(row=last, column=2, value="按功能维度查看：26项功能×（BA/开发A/开发B/测试/文档）五类角色各自的工作内容与人天，另有9类公共支撑行（初始化/架构/集成/专项/上线/管理）；底部与185人天自动对账，可与按开发顺序的「WBS分解」页对照使用")
v.alignment = Alignment(wrap_text=True, vertical="top")
ws0.row_dimensions[last].height = 30

wb.save(PATH)
print(f"完成：功能视角分解 26功能行+9公共行 | Mandy {mandy_tot}（BA {tot_cols['BA']}/测试 {tot_cols['测试']}/文档 {tot_cols['文档']}）Wade {tot_cols['Wade']} DevB {tot_cols['DevB']} PM {tot_cols['PM']} = {grand}人天 ✓")
