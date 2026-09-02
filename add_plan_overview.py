# -*- coding: utf-8 -*-
"""⑭ 计划总览：按最新「WBS分解」（V3.0）汇总生成「计划总览」页。
数据全部从 WBS分解 实时计算，并断言与 195.5 / 65 / 60 / 70.5 / 里程碑 一致。幂等可重跑。"""
import datetime as dt
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

XL = "RIMS项目WBS工作分解_AI Native版.xlsx"
HOL = {dt.date(2026, 9, 25)} | {dt.date(2026, 10, d) for d in range(1, 8)}

THIN = Side(style="thin", color="B0B0B0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
FH = Font(bold=True, size=10, color="FFFFFF")
FT = Font(bold=True, size=13, color="FFFFFF")
F1 = Font(bold=True, size=10)
F2 = Font(bold=True, size=10, color="1F4E79")
FI = Font(italic=True, size=9, color="C00000")
NAVY = PatternFill("solid", fgColor="1F4E79")
L1 = PatternFill("solid", fgColor="DDEBF7")
L2 = PatternFill("solid", fgColor="F2F2F2")
MSF = PatternFill("solid", fgColor="FFF2CC")
C = Alignment(horizontal="center", vertical="center", wrap_text=True)
L = Alignment(horizontal="left", vertical="center", wrap_text=True)

wb = openpyxl.load_workbook(XL)
ws = wb["WBS分解"]

# ---------- 从 WBS分解 提取 ----------
blocks = []  # (row_lv1, 名称, 起, 止, {角色:人天}, 任务数)
for r in range(4, 323):
    if ws.cell(r, 3).value == 1:
        blocks.append([r, None, ws.cell(r, 6).value, ws.cell(r, 7).value, {}, 0])
for i, b in enumerate(blocks):
    a = b[0]
    z = blocks[i + 1][0] - 1 if i + 1 < len(blocks) else 322
    b[1] = ws.cell(a, 2).value
    for r in range(a + 1, z + 1):
        if ws.cell(r, 3).value == 3 and isinstance(ws.cell(r, 9).value, (int, float)):
            o = ws.cell(r, 5).value
            b[4][o] = round(b[4].get(o, 0) + ws.cell(r, 9).value, 2)
            b[5] += 1
GRAND = {o: round(sum(b[4].get(o, 0) for b in blocks), 2) for o in ("Mandy", "Wade", "DevB")}
assert GRAND == {"Mandy": 65.0, "Wade": 60.0, "DevB": 70.5}, GRAND
assert round(sum(GRAND.values()), 2) == 195.5

l3 = [r for r in range(4, 323) if ws.cell(r, 3).value == 3 and ws.cell(r, 7).value]
d0 = min(ws.cell(r, 6).value for r in l3).date()
d1 = max(ws.cell(r, 7).value for r in l3).date()
assert (d0, d1) == (dt.date(2026, 9, 14), dt.date(2026, 12, 29)), (d0, d1)

def wd(a, b):
    n, d = 0, a
    while d <= b:
        if d.weekday() < 5 and d not in HOL:
            n += 1
        d += dt.timedelta(days=1)
    return n

WINDOW_WD = wd(d0, d1)
assert WINDOW_WD == 71, WINDOW_WD

last, cnt, first = {}, {}, {}
for r in l3:
    o = ws.cell(r, 5).value
    s, e = ws.cell(r, 6).value.date(), ws.cell(r, 7).value.date()
    last[o] = max(last.get(o, e), e)
    first[o] = min(first.get(o, s), s)
    cnt[o] = cnt.get(o, 0) + 1

# 里程碑（以「里程碑计划」页为准，逐条断言）
msx = wb["里程碑计划"]
MS = []
for r in range(3, 8):
    MS.append([msx.cell(r, c).value for c in range(1, 6)])
assert [m[1][:10] for m in MS] == ["2026/09/17", "2026/11/17", "2026/12/18", "2026/12/29", "2026/12/29"], MS

PHASE_DESC = [
    ("P0·第1周启动",
     "N1工程初始化与开发基座（仓库/分支/编码规范/前端骨架/后端骨架/四道门禁CI/CD）＋N4-1/2/3部署流水线与测试环境K8s先期部署",
     "M1（9/17·第1周）：流水线可一键部署更新测试环境"),
    ("P0·最高业务优先",
     "批次1：F08用户权限（角色权限/账号/SSO/身份识别/登录控制）＋F01存量迁移（结构化/非结构化导入/规则配置/迁移前后完整性）＋F02元数据（统一模型/来源绑定）",
     "M2（11/17）：批次1全部子任务测试通过（功能＋测试＋SLC）"),
    ("P1·次优先",
     "批次2：F03处置（到期删除/Legal Hold）＋F04安全（加密存储/密钥/备份）＋F05检索（关键词/系统/时间/列表/预览）＋F06审计（归档/查询/导出日志）",
     "M3（12/18）：批次2全部子任务测试通过（功能＋测试＋SLC）"),
    ("全局收口",
     "N3全局非功能：性能压测与优化/统一鉴权与敏感掩码/可观测性（日志监控告警）/容量评估，随批次并行启动、全局达标收口",
     "NFR指标达标（并入M4）"),
    ("全局收口",
     "N2代码质量治理：ESLint/SonarLint接入、SAST/SCA静态扫描随批次1启动、技术债偿还与外部评审，全部代码完成后清零",
     "M4（12/29）：静态扫描清零（与N3共同收口）"),
    ("收尾",
     "N5开发侧文档（架构/接口/数据/SLC合稿）＋N4-4/5部署预演与上线值守、运维交接",
     "M5（12/29）：文档定稿、预演通过、195.5人天全部完成"),
]
assert len(PHASE_DESC) == len(blocks) == 6

# ---------- 写「计划总览」页 ----------
if "计划总览" in wb.sheetnames:
    del wb["计划总览"]
ov = wb.create_sheet("计划总览", 1)
ov.sheet_properties.tabColor = "1F4E79"
widths = {"A": 5, "B": 26, "C": 52, "D": 11, "E": 11, "F": 8, "G": 8, "H": 8, "I": 8, "J": 8, "K": 36}
for c, w in widths.items():
    ov.column_dimensions[c].width = w

def bar(r, text):
    ov.merge_cells(start_row=r, start_column=1, end_row=r, end_column=11)
    c = ov.cell(r, 1, text)
    c.font = Font(bold=True, size=11, color="FFFFFF")
    c.fill = NAVY
    c.alignment = L
    ov.row_dimensions[r].height = 20

def kv(r, k, v, h=None):
    ov.cell(r, 2, k).font = F1
    ov.cell(r, 2).alignment = L
    ov.merge_cells(start_row=r, start_column=3, end_row=r, end_column=11)
    c = ov.cell(r, 3, v)
    c.alignment = L
    if h:
        ov.row_dimensions[r].height = h

r = 1
ov.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)
ov.cell(1, 1, "RIMS退役系统数据归档平台 · Phase1（基础归档与查询）—— 开发WBS 计划总览（V3.0）")
ov["A1"].font = FT
ov["A1"].fill = NAVY
ov["A1"].alignment = C
ov.row_dimensions[1].height = 26
r = 2
ov.merge_cells(start_row=2, start_column=1, end_row=2, end_column=11)
ov.cell(2, 1, "依据：「WBS分解」（V3.0，322行）与「功能拆解估算」——人天=拆解估算直估原值，逐字匹配，不折算、不加环节").font = Font(italic=True, size=9, color="C00000")
ov.cell(2, 1).alignment = L
r = 4
bar(r, "一、基本信息"); r += 1
INFO = [
    ("计划版本", "V3.0（开发WBS）· 编制日期 2026/9/2"),
    ("计划窗口", "2026/9/14（周一）启动 → 2026/12/29（周二）完成，共71个净工作日（77个周一~周五－6天假日；当年不跨年）"),
    ("总工作量", "195.5人天 = BA 12.5 ＋ 开发A·Wade 60 ＋ 开发B·DevB 70.5 ＋ 测试 38.75 ＋ SLC 13.75（与「功能拆解估算」完全一致）"),
    ("团队", "Wade（前端/应用）· DevB（数据/服务，关键资源：70.5人天/71工作日全占用）· Mandy（BA/测试/SLC 三职）"),
    ("排期口径", "人天=拆解估算直估原值，不折算、不加环节；无PM任务；按做项目的先后顺序与优先级排任务（框架搭建在第1周）"),
    ("日历规则", "周末不排；中秋9/25（五）、国庆10/1（四）~10/7（三）不排；调休上班日9/20（日）、10/10（六）逢周末本就不排"),
    ("计划范围", "开发WBS：25项功能＋N1~N5非功能（95子任务×角色拆行，282条任务行）；不含需求/架构设计/CC迁移演练/UAT/PM等项目环节"),
    ("排期方式", "BA澄清前置到各批次开发之前；测试随功能完成插空执行；六阶段并行交叠，里程碑定收口点"),
]
for k, v in INFO:
    kv(r, k, v, 28 if len(v) > 55 else None)
    r += 1
r += 1

bar(r, "二、阶段计划总表（六阶段·并行交叠）"); r += 1
head = ["序号", "阶段（优先级）", "主要内容与范围", "开始", "结束", "工作日", "人天", "Mandy", "Wade", "DevB", "关键交付 / 出口标准"]
for i, h in enumerate(head, 1):
    c = ov.cell(r, i, h)
    c.font = FH
    c.fill = NAVY
    c.alignment = C
    c.border = BORDER
r += 1
CN = "一二三四五六"
for i, b in enumerate(blocks):
    name = b[1].split("、", 1)[1]
    prio, desc, deliv = PHASE_DESC[i]
    a0, a1 = b[2].date(), b[3].date()
    row = [CN[i], name, desc, a0.strftime("%m/%d"), a1.strftime("%m/%d"),
           wd(a0, a1), round(sum(b[4].values()), 2), b[4].get("Mandy", 0), b[4].get("Wade", 0), b[4].get("DevB", 0), deliv]
    for cc, v in enumerate(row, 1):
        c = ov.cell(r, cc, v)
        c.border = BORDER
        c.alignment = L if cc in (2, 3, 11) else C
        if cc == 2:
            c.font = F2
        if cc == 7:
            c.font = F1
        if i in (0,):
            c.fill = L1
    ov.row_dimensions[r].height = 44
    r += 1
tot_row = ["", "合计", "六阶段并行交叠（测试插空执行，工作日列含重叠不可加总）", "09/14", "12/29", 71, 195.5, 65.0, 60.0, 70.5, "M1~M5 全部达成"]
for cc, v in enumerate(tot_row, 1):
    c = ov.cell(r, cc, v)
    c.border = BORDER
    c.fill = L2
    c.font = F1
    c.alignment = L if cc in (2, 3, 11) else C
r += 2

bar(r, "三、里程碑计划"); r += 1
mh = ["", "里程碑", "日期", "交付物 / 成果", "退出条件"]
for i, h in enumerate(mh, 1):
    if i == 1:
        continue
    c = ov.cell(r, i, h)
    c.font = FH
    c.fill = NAVY
    c.alignment = C
    c.border = BORDER
ov.merge_cells(start_row=r, start_column=4, end_row=r, end_column=7)
ov.merge_cells(start_row=r, start_column=8, end_row=r, end_column=11)
c = ov.cell(r, 4, "交付物 / 成果"); c.font = FH; c.fill = NAVY; c.alignment = C; c.border = BORDER
c = ov.cell(r, 8, "退出条件"); c.font = FH; c.fill = NAVY; c.alignment = C; c.border = BORDER
r += 1
for m in MS:
    ov.cell(r, 2, m[0]).font = F1
    ov.cell(r, 2).alignment = L
    ov.cell(r, 3, m[1][:10] + m[1][10:]).alignment = C
    ov.cell(r, 3).font = F1
    ov.merge_cells(start_row=r, start_column=4, end_row=r, end_column=7)
    ov.merge_cells(start_row=r, start_column=8, end_row=r, end_column=11)
    ov.cell(r, 4, m[3]).alignment = L
    ov.cell(r, 8, m[4]).alignment = L
    for cc in range(2, 12):
        ov.cell(r, cc).border = BORDER
    ov.row_dimensions[r].height = 30
    r += 1
r += 1

bar(r, "四、资源投入与负荷"); r += 1
rh = ["", "成员（角色）", "职责与投入区间", "任务数", "人天", "首次投入", "最后任务", "负荷说明"]
for i, h in enumerate(rh, 1):
    if i == 1:
        continue
    c = ov.cell(r, i, h)
    c.font = FH
    c.fill = NAVY
    c.alignment = C
    c.border = BORDER
ov.merge_cells(start_row=r, start_column=8, end_row=r, end_column=11)
c = ov.cell(r, 8, "负荷说明"); c.font = FH; c.fill = NAVY; c.alignment = C; c.border = BORDER
r += 1
RES = [
    ("Wade（前端/应用）", "页面/交互/应用侧开发，兼部署预演与值守",
     "60.0人天＝拆解估算原值；12/15完成全部任务，末段可支援联调", False),
    ("DevB（数据/服务）", "数据模型/迁移/服务端开发，兼测试环境与N3容量",
     "70.5人天/71工作日＝窗口内全占用（零缓冲，12/29收口）；请假/阻塞一天即顺延一天", True),
    ("Mandy（BA/测试/SLC）", "BA规则澄清前置、测试随功能完成插空执行、SLC按组汇编",
     "65.0人天（BA12.5＋测试38.75＋SLC13.75）；12/25完成", False),
]
assert GRAND["Wade"] == 60.0 and GRAND["DevB"] == 70.5 and GRAND["Mandy"] == 65.0
for name, duty, note, crit in RES:
    who = name.split("（")[0]
    ov.cell(r, 2, name).font = F1
    ov.cell(r, 2).alignment = L
    ov.cell(r, 3, duty).alignment = L
    ov.merge_cells(start_row=r, start_column=3, end_row=r, end_column=3)
    ov.cell(r, 4, cnt[who]).alignment = C
    ov.cell(r, 5, GRAND[who]).alignment = C
    ov.cell(r, 5).font = F1
    ov.cell(r, 6, first[who].strftime("%m/%d")).alignment = C
    ov.cell(r, 7, last[who].strftime("%m/%d")).alignment = C
    ov.merge_cells(start_row=r, start_column=8, end_row=r, end_column=11)
    ov.cell(r, 8, note).alignment = L
    if crit:
        for cc in range(2, 12):
            ov.cell(r, cc).fill = MSF
    for cc in range(2, 12):
        ov.cell(r, cc).border = BORDER
    ov.row_dimensions[r].height = 30
    r += 1
ov.cell(r, 2, "合计").font = F1
ov.cell(r, 3, "三人小队，无PM/环节行").alignment = L
ov.cell(r, 4, sum(cnt.values())).alignment = C
ov.cell(r, 5, 195.5).alignment = C
ov.cell(r, 5).font = F1
ov.cell(r, 6, "09/14").alignment = C
ov.cell(r, 7, "12/29").alignment = C
ov.merge_cells(start_row=r, start_column=8, end_row=r, end_column=11)
ov.cell(r, 8, "195.5人天＝BA12.5/Wade60/DevB70.5/测试38.75/SLC13.75").alignment = L
for cc in range(2, 12):
    ov.cell(r, cc).border = BORDER
    ov.cell(r, cc).fill = L2
r += 2

bar(r, "五、排期原则与阅读说明"); r += 1
NOTES = [
    "1. 人天口径：每行任务人天=「功能拆解估算」直估原值（未乘AI执行系数、未加管理环节），WBS分解末行322，层级3任务行282条合计195.5人天。",
    "2. 顺序与优先级：第1周框架搭建（N1基座＋N4流水线/测试环境先期部署）→批次1·P0（F08/F01/F02，登录与迁移主链最先可用）→批次2·P1（F03~F06）→全局非功能N3→质量治理N2→文档N5＋部署收尾。",
    "3. 并行交叠：阶段日期区间有重叠属正常——BA澄清前置、测试随功能完成插空执行、N2静态扫描随批次1启动、N3随批次2启动；里程碑（M1~M5）为各阶段收口点，详见表三。",
    "4. 关键资源：DevB 70.5人天在71个工作日内全占用（零缓冲）——决定12/29完工；Wade 61天占用（12/15完成）、Mandy 69天（12/25完成），末段可支援。",
    "5. 日历：周末与中秋9/25、国庆10/1~10/7不排；如需提前，可启用周末赶工（约可压缩2~3周）。",
    "6. 范围外：需求确认、架构评审、CC迁移演练、UAT、PM等项目环节不在本开发WBS内（195.5人天不含），需另行叠加项目级排期。",
]
for t in NOTES:
    ov.merge_cells(start_row=r, start_column=1, end_row=r, end_column=11)
    c = ov.cell(r, 1, t)
    c.alignment = L
    c.font = Font(size=9)
    ov.row_dimensions[r].height = 24
    r += 1

wb.save(XL)
print(f"计划总览完成：{r-1}行 | 阶段6 | 里程碑{len(MS)} | 总量{round(sum(GRAND.values()),2)}（Mandy {GRAND['Mandy']}/Wade {GRAND['Wade']}/DevB {GRAND['DevB']}）窗口{d0}→{d1}·{WINDOW_WD}工作日")
print("最后任务：", {k: str(v) for k, v in last.items()})
