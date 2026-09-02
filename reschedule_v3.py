# -*- coding: utf-8 -*-
"""V3.0 开发WBS：任务与人天严格=「功能拆解估算」（Wade60/DevB70.5/Mandy65=195.5），
按项目先后顺序与优先级排：第1周框架与基座(N1+N4流水线/测试环境)→批次1(F08/F01/F02)
→批次2(F03~F06)→全局非功能N3→质量治理N2→文档N5与部署收尾N4-4/5。不折算、不加环节。
用法：python3 reschedule_v3.py
"""
import ast, datetime as dt
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

PATH = "/home/user/wbs/RIMS项目WBS工作分解_AI Native版.xlsx"
src = open("/home/user/wbs/func_breakdown.py", encoding="utf-8").read()
_ns = {}
for node in ast.parse(src).body:
    if isinstance(node, ast.Assign) and getattr(node.targets[0], "id", "") in ("B", "N", "SLC_BY_GRP"):
        exec(compile(ast.Module([node], []), "<x>", "exec"), _ns)
B, N, SLC = _ns["B"], _ns["N"], _ns["SLC_BY_GRP"]
FEAT = {fc: subs for fc, subs in B}
NGRP = {gc: subs for gc, gn, subs in N}

wb = openpyxl.load_workbook(PATH)
FUNC_NAME = {}
for r in range(3, wb["26项功能覆盖对照"].max_row + 1):
    fid = wb["26项功能覆盖对照"].cell(r, 1).value
    if fid and str(fid).startswith("F"):
        FUNC_NAME[str(fid)] = str(wb["26项功能覆盖对照"].cell(r, 5).value)

HOL = {dt.date(2026, 9, 25)} | {dt.date(2026, 10, d) for d in range(1, 8)}
WD = []
_d = dt.date(2026, 9, 14)
while _d <= dt.date(2026, 12, 31):
    if _d.weekday() < 5 and _d not in HOL:
        WD.append(_d)
    _d += dt.timedelta(days=1)
WDI = {d: i for i, d in enumerate(WD)}
DEADLINE = dt.date(2026, 12, 31)

def wd_span(a, b):
    return 0 if (not a or not b or b < a) else WDI[b] - WDI[a] + 1

TASKS = {}
ORDER = {"Wade": [], "DevB": [], "Mandy": []}

def T(tid, owner, name, eff, phase, group, ai, deps=(), ms=False, deliv=None):
    assert tid not in TASKS and eff >= 0, tid
    TASKS[tid] = dict(id=tid, owner=owner, name=name, eff=float(eff), phase=phase, group=group,
                      ai=ai, deps=list(deps), ms=ms, deliv=deliv,
                      remain=float(eff), start=None, end=None, days=set())

def ai_of(name):
    if any(k in name for k in ("验证", "评审", "预演", "值守", "走查", "核对")):
        return "人工主导"
    if any(k in name for k in ("部署", "流水线", "导入作业", "备份", "监控", "【原生】")):
        return "自动化流水线"
    return "AI生成+人工审核"

BATCH1 = ["F08-01", "F08-02", "F08-03", "F08-04", "F08-05", "F01-01", "F01-02", "F01-03", "F01-04", "F01-05", "F02-01", "F02-02"]
BATCH2 = ["F03-01", "F03-02", "F04-01", "F04-02", "F04-03", "F05-01", "F05-02", "F05-03", "F05-04", "F05-05", "F06-01", "F06-02", "F06-03"]

def add_rows(gkey, subs, phase, gname, slc):
    ids = {"ba": [], "w": [], "d": [], "te": [], "slc": []}
    for i, x in enumerate(subs, 1):
        nm, desc = x[0], x[1]
        if x[2] > 0:
            tid = f"{gkey}_ba{i}"
            T(tid, "Mandy", f"{nm}（BA：规则定义与验收口径）", x[2], phase, gname, "AI辅助")
            ids["ba"].append(tid)
        if x[3] > 0:
            tid = f"{gkey}_w{i}"
            T(tid, "Wade", f"{nm}（前端/应用侧）", x[3], phase, gname, ai_of(nm))
            ids["w"].append(tid)
        if x[4] > 0:
            tid = f"{gkey}_d{i}"
            T(tid, "DevB", f"{nm}（数据/服务侧）", x[4], phase, gname, ai_of(nm))
            ids["d"].append(tid)
        if x[5] > 0:
            tid = f"{gkey}_te{i}"
            T(tid, "Mandy", f"{nm}（测试：用例与执行）", x[5], phase, gname, "AI生成+人工审核")
            ids["te"].append(tid)
    if slc > 0:
        tid = f"{gkey}_slc"
        T(tid, "Mandy", f"SLC文档汇编（{gkey}，按功能/组预估）", slc, phase, gname, "AI辅助")
        ids["slc"].append(tid)
    return ids

# ---------- P1 框架与基座（第1周）----------
G11 = "1.1 工程初始化与开发基座（N1）"
G12 = "1.2 部署流水线与测试环境先期部署（N4-1/2/3）"
n1, n4 = NGRP["N1"], NGRP["N4"]
i11 = add_rows("N1", n1, 1, G11, SLC["N1"])
i12 = add_rows("N4a", n4[:3], 1, G12, 0)
# 基座内依赖：骨架验证依赖对应骨架；CI/CD验证两侧齐；测试环境部署依赖B侧流水线
if i11["te"]:
    pass
for t in ("N1_te1",):
    if t in TASKS: TASKS[t]["deps"] = ["N1_w1", "N1_d1"]
if "N1_te2" in TASKS: TASKS["N1_te2"]["deps"] = ["N1_w2"]
if "N1_te3" in TASKS: TASKS["N1_te3"]["deps"] = ["N1_d3"]
if "N1_te5" in TASKS: TASKS["N1_te5"]["deps"] = ["N1_w5", "N1_d5"]
if i11["slc"]: TASKS[i11["slc"][0]]["deps"] = ["N1_te5"]
TASKS["N4a_d3"]["deps"] = ["N4a_d1"]          # 测试环境部署：B侧流水线就绪即可（与A侧生产流水线并行）
if "N4a_te3" in TASKS: TASKS["N4a_te3"]["deps"] = ["N4a_d3"]
TASKS["N4a_d3"]["ms"] = True
TASKS["N4a_d3"]["name"] += "◆M1 框架与基座就绪"
TASKS["N4a_d3"]["deliv"] = "测试环境（K8s）"

# ---------- P2/P3 功能批次（按优先级）----------
INFO = {}
for gi, fc in enumerate(BATCH1, 1):
    INFO[fc] = add_rows(fc, FEAT[fc], 2, f"2.{gi} {fc} {FUNC_NAME[fc]}", SLC.get(fc, 0))
for gi, fc in enumerate(BATCH2, 1):
    INFO[fc] = add_rows(fc, FEAT[fc], 3, f"3.{gi} {fc} {FUNC_NAME[fc]}", SLC.get(fc, 0))
for fc, d in INFO.items():
    dev = d["w"] + d["d"]
    for t in dev:
        TASKS[t]["deps"] = d["ba"][-1:]        # BA澄清先行
    if d["te"]:
        TASKS[d["te"][-1]]["deps"] = dev[-1:]  # 组测试在组内开发后
    if d["slc"]:
        TASKS[d["slc"][0]]["deps"] = d["te"][-1:] or dev[-1:]
TASKS[INFO["F02-02"]["te"][-1]]["ms"] = True
TASKS[INFO["F02-02"]["te"][-1]]["name"] += "◆M2 批次1完成"
TASKS[INFO["F06-03"]["te"][-1]]["ms"] = True
TASKS[INFO["F06-03"]["te"][-1]]["name"] += "◆M3 批次2完成"

# ---------- P4 全局非功能 N3 / P5 质量治理 N2 ----------
G41 = "4.1 全局非功能实现与优化（N3）"
G51 = "5.1 代码质量与静态扫描治理（N2）"
i41 = add_rows("N3", NGRP["N3"], 4, G41, SLC["N3"])
i51 = add_rows("N2", NGRP["N2"], 5, G51, SLC["N2"])
_b1w = [INFO[fc]["w"][-1] for fc in BATCH1 if INFO[fc]["w"]]
_b1d = [INFO[fc]["d"][-1] for fc in BATCH1 if INFO[fc]["d"]]
_b2w = [INFO[fc]["w"][-1] for fc in BATCH2 if INFO[fc]["w"]]
_b2d = [INFO[fc]["d"][-1] for fc in BATCH2 if INFO[fc]["d"]]
for t in i41["w"]: TASKS[t]["deps"] = _b2w[-1:]
for t in i41["d"]: TASKS[t]["deps"] = _b2d[-1:]
if i41["te"]: TASKS[i41["te"][-1]]["deps"] = [i41["w"][-1], i41["d"][0]]   # 查询性能完成即开测，容量结论随后补录
if i41["ba"]: TASKS[i41["ba"][-1]]["deps"] = ["N1_te1"]
for t in ("N2_w3", "N2_w4"):
    TASKS[t]["deps"] = _b1w[-1:]
for t in ("N2_w1", "N2_w5"):
    TASKS[t]["deps"] = _b2w[-1:]
for t in ("N2_d3", "N2_d4"):
    TASKS[t]["deps"] = _b1d[-1:]
for t in ("N2_d2", "N2_d5"):
    TASKS[t]["deps"] = _b2d[-1:]
TASKS["N2_te3"]["deps"] = ["N2_w3", "N2_d3"]
TASKS["N2_te5"]["deps"] = [i51["w"][-1], "N2_d5"]
if i51["slc"]: TASKS[i51["slc"][0]]["deps"] = i51["te"][-1:] or [i51["w"][-1]]
TASKS[i51["d"][-1]]["ms"] = True
TASKS[i51["d"][-1]]["name"] += "◆M4 质量收口"

# ---------- P6 文档与部署收尾 ----------
G61 = "6.1 开发侧文档（N5）"
G62 = "6.2 部署预演与上线值守（N4-4/5）"
i61 = add_rows("N5", NGRP["N5"], 6, G61, SLC["N5"])
i62 = add_rows("N4b", n4[3:], 6, G62, 0)
for t in i61["w"]: TASKS[t]["deps"] = [i51["w"][-1]]
for t in i61["d"]: TASKS[t]["deps"] = _b2d[-1:]
if i61["slc"]: TASKS[i61["slc"][0]]["deps"] = i61["w"][-1:] + i61["d"][-1:]
TASKS["N4b_d1"]["deps"] = ["N4a_d3"]
TASKS[i62["w"][-1]]["deps"] = [i51["w"][-1]]
TASKS[i62["w"][-1]]["ms"] = True
TASKS[i62["w"][-1]]["name"] += "◆M5"
T("N4slc", "Mandy", "SLC文档汇编（N4·部署与运维）", SLC["N4"], 6, G62, "AI辅助", deps=["N4b_d1"])

# ---------- 队列（每人严格串行；Mandy批次BA前置）----------
Wade_ORDER = ["N1_w1", "N1_w4", "N1_w2", "N1_w5", "N4a_w1", "N4a_w2"] + \
             [t for fc in BATCH1 for t in INFO[fc]["w"]] + ["N2_w3", "N2_w4"] + \
             [t for fc in BATCH2 for t in INFO[fc]["w"]] + \
             i41["w"] + ["N2_w1", "N2_w5"] + i61["w"] + i62["w"]
DevB_ORDER = ["N1_d1", "N1_d3", "N1_d4", "N1_d5", "N4a_d1", "N4a_d2", "N4a_d3"] + \
             [t for fc in BATCH1 for t in INFO[fc]["d"]] + ["N2_d3", "N2_d4"] + \
             [t for fc in BATCH2 for t in INFO[fc]["d"]] + \
             ["N2_d2", "N2_d5"] + i61["d"] + i41["d"] + i62["d"]
Mandy_ORDER = (["N1_te1"] +
               [t for fc in BATCH1 for t in INFO[fc]["ba"]] +
               [t for fc in BATCH2 for t in INFO[fc]["ba"]] +   # 全部BA前置：开发不等待澄清
               ["N1_te2", "N1_te3", "N1_te5", "N4a_te3"] +
               [t for t in i11["slc"]] +
               [t for fc in BATCH1 for t in INFO[fc]["te"] + INFO[fc]["slc"]] +
               i51["te"][:-1] + i41["ba"] + i41["te"][:-1] +
               [t for fc in BATCH2 for t in INFO[fc]["te"] + INFO[fc]["slc"]] +
               i61["slc"] + i41["te"][-1:] + i51["te"][-1:] + i51["slc"] +
               [t for t in i62["te"]] + ["N4slc"])
for who, order in (("Wade", Wade_ORDER), ("DevB", DevB_ORDER), ("Mandy", Mandy_ORDER)):
    ORDER[who] = order
    have = {t for t, v in TASKS.items() if v["owner"] == who}
    assert set(order) == have, f"{who}队列差: 缺{have - set(order)} 多{set(order) - have}"

# ---------- 调度 ----------
cur = 0
cursors = {w: 0 for w in ORDER}
while any(cursors[w] < len(ORDER[w]) for w in ORDER):
    assert cur < len(WD), "日历耗尽"
    day = WD[cur]
    for w, order in ORDER.items():
        cap = 1.0
        # 插空式：优先续做已开工任务，再从队首起找依赖就绪的任务开工（不改变人天与依赖）
        for pass_cont in (True, False):
            if cap <= 1e-9:
                break
            for tid in order[cursors[w]:cursors[w] + 16]:   # 插空窗口：只向后看16个任务
                t = TASKS[tid]
                if t["remain"] <= 1e-9:
                    continue
                if (t["start"] is not None) != pass_cont:
                    continue
                if any(TASKS[d]["end"] is None or TASKS[d]["end"] >= day for d in t["deps"]):
                    continue
                t["start"] = t["start"] or day
                take = min(cap, t["remain"])
                t["remain"] -= take; cap -= take; t["days"].add(day)
                if t["remain"] <= 1e-9:
                    t["end"] = day
            while cursors[w] < len(order) and TASKS[order[cursors[w]]]["remain"] <= 1e-9:
                cursors[w] += 1
    cur += 1
_late = [(t["id"], t["end"]) for t in TASKS.values() if t["end"] and t["end"] > DEADLINE]
assert not _late, f"超出12/31: {_late[:8]}"
_tot = {}
for t in TASKS.values():
    _tot[t["owner"]] = round(_tot.get(t["owner"], 0) + t["eff"], 2)
assert _tot == {"Wade": 60.0, "DevB": 70.5, "Mandy": 65.0}, _tot
M1 = TASKS["N4a_d3"]["end"]; M2 = TASKS[INFO["F02-02"]["te"][-1]]["end"]
M3 = TASKS[INFO["F06-03"]["te"][-1]]["end"]; M4 = max(TASKS[i51["d"][-1]]["end"], TASKS[i41["d"][-1]]["end"])
M5 = max(t["end"] for t in TASKS.values() if t["eff"] > 0)

# ---------- 渲染 ----------
ws = wb["WBS分解"]
wb.remove(ws)
ws = wb.create_sheet("WBS分解", 1)
THIN = Side(style="thin", color="B0B0B0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
FN = Font(bold=True, size=13, color="FFFFFF"); FH = Font(bold=True, size=10, color="FFFFFF")
F1 = Font(bold=True, size=10); F2 = Font(bold=True, size=10, color="1F4E79"); FM = Font(bold=True, size=10, color="9C5700")
NAVY = PatternFill("solid", fgColor="1F4E79"); L1 = PatternFill("solid", fgColor="DDEBF7")
L2 = PatternFill("solid", fgColor="F2F2F2"); MSF = PatternFill("solid", fgColor="FFF2CC")
C = Alignment(horizontal="center", vertical="center", wrap_text=True)
L = Alignment(horizontal="left", vertical="center", wrap_text=True)
for c, w in {"A": 10, "B": 54, "C": 6, "D": 26, "E": 11, "F": 11, "G": 11, "H": 9, "I": 11, "J": 15, "K": 16, "L": 9, "M": 8, "N": 8, "O": 11, "P": 11, "Q": 40}.items():
    ws.column_dimensions[c].width = w
ws.merge_cells("A1:Q1")
ws["A1"] = "RIMS数据归档平台Phase1 · 开发WBS — V3.0严格按「功能拆解估算」排期：BA12.5/Wade60/DevB70.5/测试38.75/SLC13.75＝195.5人天（先框架后功能、按优先级分批）"
ws["A1"].font = FN; ws["A1"].fill = NAVY; ws["A1"].alignment = C
ws.row_dimensions[1].height = 24
ws["A2"] = "项目名称：RIMS退役系统数据归档平台·Phase1（基础归档与查询）"
ws["F2"] = "开发团队：Wade（前端/应用）·DevB（数据/服务）·Mandy（BA/测试/SLC）"
ws["I2"] = f"编制日期：2026/9/2  版本：V3.0  窗口：2026/9/14 → {M5.strftime('%Y/%m/%d')}（第1周框架搭建；人天=拆解估算直估原值）"
HD = ["WBS编号", "任务名称", "层级", "交付物/成果", "负责人", "开始日期", "结束日期", "工期(天)", "工作量(人天)", "AI参与方式", "前置任务", "状态", "进度", "优先级", "预算(元)", "实际成本(元)", "备注"]
for i, h in enumerate(HD, 1):
    c = ws.cell(3, i, h); c.font = FH; c.fill = NAVY; c.alignment = C; c.border = BORDER
ws.freeze_panes = "C4"

PH = {1: "一、工程基座与部署流水线（第1周·框架搭建）",
      2: "二、批次1·P0优先级：用户权限F08＋存量迁移F01＋元数据F02",
      3: "三、批次2·P1优先级：处置F03＋安全F04＋检索F05＋审计F06",
      4: "四、全局非功能实现与优化（N3）",
      5: "五、代码质量与静态扫描治理（N2）",
      6: "六、开发侧文档与部署收尾（N5＋N4-4/5）"}
GRP = {}
for ph in range(1, 7):
    seen = []
    for who in ("Mandy", "Wade", "DevB"):
        for tid in ORDER[who]:
            g = TASKS[tid]["group"]
            if TASKS[tid]["phase"] == ph and g not in seen:
                seen.append(g)
    def _key(g):
        head = g.split(" ")[0]
        return (1, [int(x) for x in head.split(".")]) if head.replace(".", "").isdigit() else (0, [0])
    GRP[ph] = sorted(seen, key=_key)
gleaves = {ph: {g: [] for g in GRP[ph]} for ph in range(1, 7)}
for who in ("Mandy", "Wade", "DevB"):
    for tid in ORDER[who]:
        t = TASKS[tid]
        gleaves[t["phase"]][t["group"]].append(tid)

row = 4
code_map = {}
per = {}

def style(r, lv, ms=False):
    for cc in range(1, 18):
        cell = ws.cell(r, cc); cell.border = BORDER
        if lv == 1: cell.font = F1; cell.fill = L1
        elif lv == 2: cell.font = F2; cell.fill = L2
        elif ms: cell.font = FM; cell.fill = MSF
        cell.alignment = L if cc in (2, 4, 10, 11, 17) else C
    for col in (6, 7):
        ws.cell(r, col).number_format = "mm/dd"
    ws.cell(r, 8).number_format = "0"; ws.cell(r, 9).number_format = "0.##"

for ph in range(1, 7):
    prow = row
    ws.cell(row, 1, str(ph)); ws.cell(row, 2, PH[ph]); ws.cell(row, 3, 1)
    style(row, 1); row += 1
    pfirst = row
    for gname in GRP[ph]:
        grow = row; row += 1
        leaves = gleaves[ph][gname]
        head = gname.split(" ")[0]
        gcode = head if head.replace(".", "").isdigit() else f"{ph}.{GRP[ph].index(gname) + 1}"
        label = gname.split(" ", 1)[1] if head.replace(".", "").isdigit() else gname
        for n, tid in enumerate(leaves, 1):
            t = TASKS[tid]
            lc = f"{gcode}.{n}"
            ws.cell(row, 1, lc); ws.cell(row, 2, t["name"]); ws.cell(row, 3, 3)
            ws.cell(row, 4, t["deliv"]); ws.cell(row, 5, t["owner"])
            ws.cell(row, 6, t["start"]); ws.cell(row, 7, t["end"])
            ws.cell(row, 8, wd_span(t["start"], t["end"])); ws.cell(row, 9, t["eff"])
            ws.cell(row, 10, t["ai"])
            if t["deps"]:
                ws.cell(row, 11, ",".join(code_map.get(d, d) for d in t["deps"])[:200])
            style(row, 3, t["ms"])
            code_map[tid] = lc
            per[t["owner"]] = round(per.get(t["owner"], 0) + t["eff"], 2)
            row += 1
        ss = [TASKS[t]["start"] for t in leaves]; ee = [TASKS[t]["end"] for t in leaves]
        ws.cell(grow, 1, gcode); ws.cell(grow, 2, label); ws.cell(grow, 3, 2)
        ws.cell(grow, 5, "/".join(sorted({TASKS[t]["owner"] for t in leaves})))
        ws.cell(grow, 6, min(ss)); ws.cell(grow, 7, max(ee))
        ws.cell(grow, 8, wd_span(min(ss), max(ee)))
        ws.cell(grow, 9, f'=ROUND(SUMPRODUCT(($C${grow+1}:$C${row-1}=3)*(LEFT($A${grow+1}:$A${row-1},LEN(A{grow})+1)=A{grow}&".")*$I${grow+1}:$I${row-1}),2)')
        style(grow, 2)
    allv = [t for g in GRP[ph] for t in gleaves[ph][g]]
    ss = [TASKS[t]["start"] for t in allv]; ee = [TASKS[t]["end"] for t in allv]
    ws.cell(prow, 5, "/".join(sorted({TASKS[t]["owner"] for t in allv})))
    ws.cell(prow, 6, min(ss)); ws.cell(prow, 7, max(ee))
    ws.cell(prow, 8, wd_span(min(ss), max(ee)))
    ws.cell(prow, 9, f'=ROUND(SUMPRODUCT(($C${pfirst}:$C${row-1}=3)*(LEFT($A${pfirst}:$A${row-1},LEN(A{prow})+1)=A{prow}&".")*$I${pfirst}:$I${row-1}),2)')
    style(prow, 1)
LAST = row - 1

# ---------- 里程碑计划 ----------
WK = {0: "一", 1: "二", 2: "三", 3: "四", 4: "五"}
msx = wb["里程碑计划"]
for mr in list(msx.merged_cells.ranges):
    msx.unmerge_cells(str(mr))
for r in range(1, msx.max_row + 1):
    for c in range(1, 9):
        msx.cell(r, c).value = None
msx["A1"] = f"开发里程碑（V3.0·拆解估算口径195.5人天 · {M5.strftime('%Y/%m/%d')}前完成）"
msx["A1"].font = Font(bold=True, size=12, color="FFFFFF"); msx["A1"].fill = NAVY; msx["A1"].alignment = C
for i, h in enumerate(["里程碑", "日期", "星期", "交付物/成果", "退出条件", "关联任务"], 1):
    c = msx.cell(2, i, h); c.font = FH; c.fill = NAVY; c.alignment = C; c.border = BORDER
MDEF = [("M1 框架与基座就绪（第1周）", M1, "CI/CD四道门禁＋部署流水线＋测试环境（K8s）先期部署", "流水线可一键部署更新测试环境", "N4a_d3"),
        ("M2 批次1完成（F08/F01/F02）", M2, "登录权限/存量迁移/元数据：功能＋测试＋SLC", "批次1全部子任务测试通过", INFO["F02-02"]["te"][-1]),
        ("M3 批次2完成（F03~F06）", M3, "处置/安全/检索/审计：功能＋测试＋SLC", "批次2全部子任务测试通过", INFO["F06-03"]["te"][-1]),
        ("M4 全局非功能与质量收口（N3＋N2）", M4, "性能/鉴权掩码/可观测/容量＋ESLint/SonarLint/SAST/SCA清零", "NFR指标达标、静态扫描清零", i51["d"][-1]),
        ("M5 开发WBS完成（文档定稿＋部署收尾）", M5, "架构/接口/数据文档＋部署预演＋运维交接＋SLC合稿", "全部195.5人天任务完成、文档定稿、预演通过", "N4slc")]
r = 3
for name, d, deliv, exit_, tid in MDEF:
    msx.cell(r, 1, name).font = Font(bold=True)
    msx.cell(r, 2, d.strftime("%Y/%m/%d") + f"（周{WK[d.weekday()]}）")
    msx.cell(r, 3, f"周{WK[d.weekday()]}")
    msx.cell(r, 4, deliv); msx.cell(r, 5, exit_); msx.cell(r, 6, code_map.get(tid, ""))
    for c in range(1, 7):
        msx.cell(r, c).border = BORDER
        msx.cell(r, c).alignment = Alignment(wrap_text=True, vertical="center", horizontal="left" if c in (4, 5) else "center")
    msx.row_dimensions[r].height = 28
    r += 1
msx.cell(r + 1, 1, "注：V3.0为开发WBS——任务与人天严格取自「功能拆解估算」（195.5人天），不含需求/架构设计/CC迁移演练/UAT/PM等项目环节；"
                   "按先后顺序与优先级排：第1周框架搭建→批次1（P0）→批次2（P1）→全局非功能→质量治理→文档与部署收尾。").font = Font(italic=True, size=9, color="C00000")

# ---------- WBS词典 ----------
d = wb["WBS词典"]
for mr in list(d.merged_cells.ranges):
    d.unmerge_cells(str(mr))
for r in range(1, d.max_row + 1):
    for c in range(1, 9):
        d.cell(r, c).value = None
d["A1"] = "WBS词典（V3.0开发WBS·关键工作包）"; d["A1"].font = Font(bold=True, size=12); d.merge_cells("A1:G1")
for i, h in enumerate(["WBS编号", "任务名称", "工作包定义", "交付物", "验收标准", "负责人"], 1):
    c = d.cell(2, i, h); c.font = FH; c.fill = NAVY; c.alignment = C; c.border = BORDER
DR = [("1.1", "四道安全门禁CI/CD搭建", "Azure DevOps：构建＋镜像＋SAST/依赖/密钥扫描＋越权用例门禁，任何合入必须过门禁。", "CI/CD流水线", "拦截样例演示；门禁不可bypass", "Wade"),
      ("1.2", "测试环境先期部署◆M1", "ADO Service Connection连接K8s（命名空间RBAC）；测试环境早期先部署、随迭代流水线自动更新；生产环境终期部署带审批；不用Helm。", "测试环境（K8s）", "流水线一键部署更新", "DevB"),
      ("2.6", "SeaTunnel批量导入作业（F01-01）", "作业配置生成器、分片/断点续传/限速执行。", "导入作业", "断点续传与限速演示通过", "DevB"),
      ("3.1", "到期自动删除处置引擎（F03-01）", "到期扫描、豁免检查（Legal Hold）、DROP PARTITION、VACUUM物理清除。", "处置引擎", "处置链路（扫描→删除→留痕）演示通过", "DevB"),
      ("3.10", "原始文件在线预览（F05-05）", "格式识别、图片/PDF/Office转换、SAS直连流式加载。", "预览服务", "多格式预览与失败降级下载", "Wade/DevB"),
      ("4.1", "可观测性【原生】", "App Insights/Log Analytics/Azure Monitor：SDK插桩、作业监控对接与仪表盘。", "监控仪表盘", "关键作业与API告警就绪", "Wade/DevB"),
      ("5.1", "SAST安全告警修复（N2）", "硬编码密钥/注入/越权路径等静态安全缺陷修复清零。", "扫描清零报告", "四道门禁零严重告警", "Wade＋DevB"),
      ("6.1", "开发侧文档（N5）", "架构与总体设计/应用接口（OpenAPI）/数据设计与ETL作业文档。", "文档V1.0", "评审通过并归档", "Wade＋DevB"),
      ("6.2", "部署预演与上线值守（N4-4/5）", "预生产完整演练、回退验证；上线窗口值守、运维交接与培训材料。", "预演记录/Runbook", "回退验证通过", "Wade/DevB")]
r = 3
for row_ in DR:
    for c, v in enumerate(row_, 1):
        cell = d.cell(r, c, v); cell.border = BORDER
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="left" if c in (3, 4, 5) else "center")
    d.row_dimensions[r].height = 34
    r += 1

# ---------- 覆盖对照编码 ----------
wc = wb["26项功能覆盖对照"]
for r in range(3, wc.max_row + 1):
    fid = wc.cell(r, 1).value
    if fid and str(fid).startswith("F"):
        fc = str(fid)
        grp = next((g for g in GRP[2] + GRP[3] if fc in g), None)
        if grp:
            head = grp.split(" ")[0]
            wc.cell(r, 7, head)
            wc.cell(r, 8, "同组·前端/后端行")
            wc.cell(r, 9, "同组·测试/SLC行")
            wc.cell(r, 10, "—（测试随功能组行）")
            ends = [TASKS[t]["end"] for t in gleaves[2].get(grp, []) + gleaves[3].get(grp, [])]
            wc.cell(r, 11, ("Wade/DevB / " + max(ends).strftime("%m/%d")) if ends else "—")

wb.save(PATH)
print("V3.0 完成 | 里程碑:", M1, M2, M3, M4, M5)
print("每人直估:", per, "| 合计:", round(sum(per.values()), 2), "| WBS末行:", LAST)
